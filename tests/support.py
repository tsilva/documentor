from pathlib import Path

import fitz
import openpyxl
from PIL import Image

from papertrail.config import ProfileSettings
from papertrail.commands.reconcile import _reconciliation_rules
from papertrail.runtime import Runtime, runtime_from_profile


TEST_COUNTERPARTY_ALIASES = {
    "TESTSHAREDTOLL": "shared-toll",
    "sharedtoll": "shared-toll",
    "sharedtollprovider": "shared-toll",
    "TESTCOMPANY": "example-company",
    "examplecompany": "example-company",
    "examplecompanyunipessoallda": "example-company",
    "benefitsprovider": "benefits-provider",
    "bankalpha": "bank-alpha",
    "bankalphasa": "bank-alpha",
    "bankbeta": "bank-beta",
    "bankbetasa": "bank-beta",
}

TEST_LINE_ITEM_EXTRACTORS = {
    "bpi_fee_invoice": {
        "document_types": ["invoice"],
        "issuing_parties": ["bpi", "bank-beta"],
        "title_terms": ["comissoes", "titulos", "manutencaodecontavalornegocios", "contavalornegocios"],
        "maintenance_marker": "MANUTENCAO DE CONTA VALOR NEGOCIOS",
        "custody_marker": "COMISSAO DEPOSITO E REGISTO VALORES MOBILIARIOS",
        "total_debit_marker": "TOTAL A DEBITO",
        "stamp_duty_rate": 0.04,
        "max_stamp_duty": 1.0,
        "maintenance_search_after": 25,
        "custody_total_search_after": 20,
        "custody_total_amount_after": 3,
        "custody_fallback_after": 15,
        "custody_fallback_before": 15,
        "custody_fallback_max_amount": 100,
        "maintenance_category": "bank-fee-maintenance",
        "stamp_duty_category": "bank-fee-stamp-duty",
        "custody_category": "bank-custody-fee",
    },
    "bpi_stock_invoice": {
        "document_types": ["invoice"],
        "issuing_parties": ["bpi", "bank-beta"],
        "title_terms": ["comissoes", "titulos"],
        "sale_required_terms": ["VENDA DE", "ACCOES"],
        "order_reference_pattern": r"N[ºO]\s+ORDEM:\s*([A-Z0-9]+)",
        "total_credit_marker": "TOTAL A CREDITO",
        "settlement_offset_days": 1,
        "movement_date_lookback": 25,
        "reference_search_after": 15,
        "category": "stock-sale-bpi",
        "document_type": "bank-stock-sell",
    },
    "bpi_transfer": {
        "document_types": ["bank-note", "bank-transfer"],
        "issuing_parties": ["bpi", "bank-beta"],
        "line_pattern": r"\b(?:TRF\s+CR\s+SEPA\+\s+|TRF\s+SEPA\+\s+INST\s+|TRANSFER[ÊE]NCIA\s+RECEBIDA\s+)(\d+)\b",
        "amount_search_before": 8,
        "date_search_after": 5,
        "category": "bank-transfer-sepa",
    },
    "millennium_fee_invoice": {
        "document_types": ["invoice", "invoice-receipt"],
        "issuing_parties": ["millenniumbcp", "millennium-bcp", "bank-alpha"],
        "movement_date_marker": "DATA DO MOVIMENTO",
        "amount_search_after": 5,
        "markers": {
            "CUSTO DE SERVICO INTERNACIONAL": "bank-fee-international-service",
            "COMISSAO REFERENTE": "bank-fee-maintenance",
        },
        "stamp_duty_markers": ["IMPOSTO DO SELO", "IMP. SELO"],
        "stamp_duty_legal_reference": "17.3.4",
        "stamp_duty_category": "bank-fee-stamp-duty",
    },
    "direct_debit": {
        "date_pattern": r"\bDEBITO\s+A\s+PARTIR\s+DE:\s*(\d{2})[/-](\d{2})[/-](20\d{2})\b",
        "auth_pattern": r"\bN[ºO]\s+AUTORIZACAO:\s*([A-Z0-9]+)\b",
        "amount_pattern": r"\bVALOR:[^\d\n]*([\d\s.]+,\d{2})\b",
        "category": "supplier-payment",
        "label": "Direct debit",
    },
    "insurance_notice": {
        "document_types": ["insurance-notice"],
        "period_pattern": r"\bPERIODO\s+DO\s+RECIBO\b.*?(\d{2})[/-](\d{2})[/-](20\d{2})\s+A\s+\d{2}[/-]\d{2}[/-]20\d{2}\b",
        "reference_pattern": r"\bADC\s+([A-Z0-9]+)\b",
        "category": "supplier-payment",
        "label": "Insurance direct debit",
    },
}

_GENERIC_RULES = [rule.model_dump() for rule in _reconciliation_rules()]
TEST_RECONCILIATION_RULES = [
    {
        "name": "investment",
        "match_description": ["TRANSFERENCIA A CREDITO LIS"],
        "required_types": {"bank-stock-sell": [1, None]},
        "expected_page_count": {"bank-stock-sell": [1, 2]},
    },
    {
        "name": "investment",
        "match_description": ["TRANSFERENCIA A DEBITO LIS"],
        "required_types": {"investment-evidence": [1, None]},
        "shared_types": {"investment-evidence": "bank-beta"},
        "shared_filters": {"investment-evidence": {"document_type": "bank-investment"}},
        "expected_page_count": {"investment-evidence": [1, 2]},
    },
    {
        "name": "loan-payment",
        "match_description": ["CONCESS CRED EMPR"],
        "required_types": {"bank-anchor": [1, None], "contract-evidence": [1, None]},
        "expected_page_count": {"bank-anchor": 1},
    },
    {
        "name": "loan-payment",
        "match_description": ["IMP ABERT CRED EMPRES"],
        "required_types": {"bank-anchor": 1, "supplier-evidence": 1},
        "expected_page_count": {"bank-anchor": 1, "supplier-evidence": 1},
    },
    {
        "name": "loan-payment",
        "match_description": ["PAGAMENT EMPRESTIMO"],
        "required_types": {"bank-anchor": 1, "supplier-evidence": 1},
        "shared_types": {"supplier-evidence": "bank-alpha"},
        "shared_filters": {"supplier-evidence": {"document_title": "Pagamento de Prestação"}},
        "expected_page_count": {"bank-anchor": 1, "supplier-evidence": 1},
    },
    {
        "name": "tax-payment",
        "match_description": ["IGCP", "PAG.DUC"],
        "required_types": {"bank-anchor": 1, "tax-evidence": 1},
        "expected_page_count": {"bank-anchor": 1},
    },
    {
        "name": "payroll-payment",
        "match_description": ["EMPLOYEE ONE", "EMPLOYEE TWO"],
        "required_types": {
            "bank-anchor": 1,
            "payroll-evidence": 1,
        },
        "shared_types": {"payroll-evidence": None},
        "shared_filters": {
            "payroll-evidence": {
                "document_type": "payroll-salary",
            },
        },
        "expected_page_count": {"bank-anchor": 1},
    },
    {
        "name": "payroll-payment",
        "match_description": ["TAXA SOCIAL UNICA"],
        "required_types": {"bank-anchor": 1, "payroll-evidence": 1},
        "expected_page_count": {"bank-anchor": 1},
    },
    {
        "name": "bank-fee",
        "match_description": [
            "CUSTO DE SERVICO INTERNACIONAL",
            "MANUTENCAO DE CONTA VALOR NEGOCIOS",
            "PACOTE M EMPRESA",
            "TITULOS CUSTODIA",
            "OPERACOES COM TITULOS",
            "IMPOSTO DO SELO",
            "IMPOSTO SELO",
            "IMPOSTO DE SELO",
            "COMISSAO",
            "COMISSÃO",
        ],
        "required_types": {"supplier-evidence": [1, None], "bank-anchor": [0, 1]},
        "expected_page_count": {"supplier-evidence": [1, 2], "bank-anchor": 1},
    },
    {
        "name": "bank-only",
        "match_description": [
            "TRF P/ EXAMPLE COMPANY - BANK BETA",
            "TRF P/ EXAMPLE COMPANY - BPI",
            "EXAMPLE COMPANY - BANK BETA",
            "EXAMPLE COMPANY - BPI",
        ],
        "required_types": {"bank-anchor": 1},
        "expected_page_count": {"bank-anchor": 1},
    },
    {
        "name": "bank-only",
        "match_description": ["TRF SEPA+"],
        "required_types": {"bank-anchor": [1, None]},
        "expected_page_count": {},
    },
    {
        "name": "bank-only",
        "match_description": ["TRANSFERENCIA RECEBIDA", "TRANSFERÊNCIA RECEBIDA"],
        "required_types": {"bank-anchor": 1},
        "expected_page_count": {"bank-anchor": 1},
    },
    {
        "name": "bank-only",
        "direction": "credit",
        "required_types": {"bank-anchor": [1, None]},
        "expected_page_count": {"bank-anchor": 1},
    },
    *_GENERIC_RULES,
]


def make_test_runtime(root: Path) -> Runtime:
    raw = root / "raw"
    processed = root / "processed"
    export = root / "export"
    cache = root / "cache"
    raw.mkdir()
    processed.mkdir()
    export.mkdir()
    cache.mkdir()

    profile = ProfileSettings.model_validate(
        {
            "profile": {"name": "test", "description": "", "tax_number": None},
            "paths": {
                "raw": [str(raw)],
                "processed": str(processed),
                "export": str(export),
            },
            "openrouter": {
                "model_id": "test-model",
                "base_url": "https://example.invalid/api/v1",
                "api_key": "test-key",
            },
            "gmail": {"enabled": False},
            "passwords": {},
            "nif_api": {"enabled": False},
            "reconciliation": {"exclude_prefixes": []},
            "export": {
                "file_mappings": {
                    "enabled": False,
                    "default_prefix": "",
                    "rules": [],
                    "filename_fields": [],
                },
                "merge_rules": [],
                "max_file_size_mb": None,
            },
            "profile_path": root / "profile.yaml",
            "profile_dir": root,
        }
    )
    profile.reconciliation.counterparty_aliases = TEST_COUNTERPARTY_ALIASES
    profile.reconciliation.statement_bank_issuer_aliases = {
        "bankalpha": "bank-alpha",
        "bankalphasa": "bank-alpha",
        "bankbeta": "bank-beta",
        "bankbetasa": "bank-beta",
        "millenniumbcp": "bank-alpha",
        "bpi": "bank-beta",
    }
    profile.reconciliation.bank_counterparties = ["bank-alpha", "bank-beta"]
    profile.reconciliation.strict_statement_banks = ["bank-beta"]
    profile.reconciliation.supporting_pair_exempt_statement_banks = ["bank-beta"]
    profile.reconciliation.shared_period_transaction_keywords = {"shared-toll": ["SHAREDTOLL"]}
    profile.reconciliation.shared_period_title_terms = {
        "shared-toll": ["sharedperiod", "monthlybundle", "pagamentosdeservicos", "extratorecibo"],
        "$bank": ["comissoes", "manctapacote", "operacaocartoes", "impostodoselo"],
    }
    profile.reconciliation.same_month_shared_rule_names = ["vendor-shared-toll"]
    profile.reconciliation.line_item_extractors = TEST_LINE_ITEM_EXTRACTORS
    profile.reconciliation.rules = TEST_RECONCILIATION_RULES

    return runtime_from_profile(
        profile,
        profiles_dir=root,
        enable_client=False,
        probe_api=False,
    )


def create_pdf(pdf_path: Path, page_texts: list[str]) -> None:
    doc = fitz.open()
    for text in page_texts:
        page = doc.new_page()
        page.insert_text((72, 72), text)
    doc.save(pdf_path)
    doc.close()


def create_png(image_path: Path, size: tuple[int, int] = (40, 40), color: str = "white") -> None:
    Image.new("RGB", size, color=color).save(image_path)


def create_millennium_statement(
    xlsx_path: Path,
    account: str = "TEST-ACCOUNT-ALPHA",
    currency: str = "EUR",
    period_start: str = "01/01/2026",
    period_end: str = "31/01/2026",
    transactions: list[dict] | None = None,
) -> None:
    transactions = transactions or [
        {
            "date_posting": "01/01/2026",
            "date_value": "01/01/2026",
            "description": "TEST PURCHASE",
            "amount": -12.34,
            "currency": currency,
            "notes": "",
            "treated": "Nao",
        }
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=3, value=f"{account} - {currency}")
    ws.cell(row=3, column=3, value=period_start)
    ws.cell(row=4, column=3, value=period_end)
    ws.cell(row=8, column=1, value="Data Lancamento")
    ws.cell(row=8, column=2, value="Data Valor")
    ws.cell(row=8, column=3, value="Descricao")
    ws.cell(row=8, column=4, value="Montante")
    ws.cell(row=8, column=5, value="Moeda")
    ws.cell(row=8, column=6, value="Notas")
    ws.cell(row=8, column=7, value="Tratado")

    for index, txn in enumerate(transactions, start=9):
        ws.cell(row=index, column=1, value=txn["date_posting"])
        ws.cell(row=index, column=2, value=txn["date_value"])
        ws.cell(row=index, column=3, value=txn["description"])
        ws.cell(row=index, column=4, value=txn["amount"])
        ws.cell(row=index, column=5, value=txn.get("currency", currency))
        ws.cell(row=index, column=6, value=txn.get("notes", ""))
        ws.cell(row=index, column=7, value=txn.get("treated", "Nao"))

    wb.save(xlsx_path)
    wb.close()


def create_bpi_statement(
    xlsx_path: Path,
    account: str = "TEST-ACCOUNT-BETA",
    currency: str = "EUR",
    transactions: list[dict] | None = None,
) -> None:
    transactions = transactions or [
        {
            "date_posting": "01-01-2026",
            "date_value": "01-01-2026",
            "description": "TEST TRANSFER",
            "amount": 12.34,
            "currency": currency,
        }
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=7, column=3, value=f"{account} ({currency})")
    ws.cell(row=18, column=1, value="Data Mov.")
    ws.cell(row=18, column=2, value="Data Valor")
    ws.cell(row=18, column=3, value="Descricao do Movimento")
    ws.cell(row=18, column=4, value="Valor em EUR")

    for index, txn in enumerate(transactions, start=19):
        ws.cell(row=index, column=1, value=txn["date_posting"])
        ws.cell(row=index, column=2, value=txn["date_value"])
        ws.cell(row=index, column=3, value=txn["description"])
        ws.cell(row=index, column=4, value=txn["amount"])

    wb.save(xlsx_path)
    wb.close()
