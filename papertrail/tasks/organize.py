"""Organize phase: renaming, export, archive, gmail download, merge attachments."""

import hashlib
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Optional

import pandas as pd

from papertrail.config import get_current_profile
from papertrail.console import get_console
from papertrail.hashing import hash_file_fast, hash_file_text
from papertrail.logging_utils import get_logger, log_failure, setup_task_logging, DocumentLogger
from papertrail.metadata import (
    find_companion_file, get_unique_dates, iter_json_files,
    load_json_data, save_json_data, save_metadata_json,
)
from papertrail.models import DocumentMetadata, clean_enum_string
from papertrail.tasks import task_log_context

logger = get_logger('cli')


# --- Filename helpers ---

def sanitize_filename_component(s: str) -> str:
    """Sanitize a string for use in a filename."""
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[\\/*?:"<>|()\[\],]', '', s).strip()
    return re.sub(r'\s+', ' ', s)


def file_name_from_metadata(metadata: DocumentMetadata, file_hash: str) -> str:
    """Generate a filename from metadata."""
    parts = [
        sanitize_filename_component(metadata.date_issued),
        sanitize_filename_component(metadata.document_type),
        sanitize_filename_component(metadata.issuing_party)
    ]

    if metadata.document_title:
        title = sanitize_filename_component(metadata.document_title)
        if len(title) > 80:
            title = title[:80].rsplit(" ", 1)[0]
        parts.append(title)

    if metadata.total_amount is not None:
        amount = f"{metadata.total_amount:.0f}" if metadata.total_amount.is_integer() else f"{metadata.total_amount:.2f}"
        currency = metadata.total_amount_currency or ""
        parts.append(sanitize_filename_component(f"{amount} {currency}".strip()))

    ext = getattr(metadata, 'source_extension', None) or '.pdf'
    parts.append(f"{file_hash}{ext}")
    return " - ".join(parts).lower()


# --- Rename ---

def rename_single_pdf(pdf_path: Path, content_hash: str, processed_path: Path,
                      known_content_hashes: set, known_file_hashes: set, failure_logger=None,
                      doc_logger: DocumentLogger = None):
    """Process and rename a single PDF file."""
    from papertrail.tasks.extraction import classify_pdf_document

    try:
        file_hash = hash_file_fast(pdf_path)
        metadata = classify_pdf_document(pdf_path, content_hash, failure_logger, doc_logger=doc_logger)
        metadata.hash_file = file_hash
        metadata.hash_text = hash_file_text(pdf_path)
        metadata.file_size_kb = round(pdf_path.stat().st_size / 1024)

        filename = file_name_from_metadata(metadata, file_hash)
        new_pdf_path = processed_path / filename

        if new_pdf_path.exists():
            logger.warning(f"Skipping {pdf_path.name}: destination already exists: {filename}")
            return

        shutil.copy2(pdf_path, new_pdf_path)
        save_metadata_json(new_pdf_path, metadata)

        known_content_hashes.add(content_hash)
        known_file_hashes.add(file_hash)
        logger.debug(f"Processed: {pdf_path.name} -> {filename}")
    except Exception as e:
        log_failure(failure_logger, pdf_path, e)
        logger.error(f"Failed to process {pdf_path.name}: {e}")


def rename_pdf_files(pdf_paths, file_hash_map, known_content_hashes, known_file_hashes, processed_path,
                     failure_logger=None, doc_logger: DocumentLogger = None):
    """Rename multiple PDF files."""
    console = get_console()

    with console.progress("Extracting", total=len(pdf_paths)) as progress:
        task = progress.add_task("Extracting", total=len(pdf_paths))
        for pdf_path in pdf_paths:
            name = pdf_path.stem
            if len(name) > 40:
                name = name[:37] + "..."
            progress.update(task, description=f"[dim]{name}[/dim]")
            rename_single_pdf(pdf_path, file_hash_map[pdf_path], processed_path, known_content_hashes, known_file_hashes,
                              failure_logger, doc_logger=doc_logger)
            progress.update(task, advance=1)


def task_rename_files(processed_path: Path, quiet: bool = False) -> dict:
    """Rename existing PDF files based on metadata."""
    from papertrail.metadata import load_json_files_parallel

    console = get_console()

    with task_log_context(processed_path, "rename_files", show_header=not quiet):
        logger.debug("Renaming existing PDF files and metadata based on metadata...")

        valid_entries = []
        orphan_count = 0

        for metadata_path, metadata in load_json_files_parallel(processed_path, validate=True, show_progress=not quiet, progress_desc="Validating metadata"):
            doc_path = find_companion_file(metadata_path, metadata.model_dump())
            if doc_path is None:
                orphan_count += 1
                logger.warning(f"Skipping {metadata_path.name}: companion file not found")
                continue

            valid_entries.append((doc_path, metadata))

        logger.debug(f"Found {len(valid_entries)} files to validate")

        renamed_count = 0
        for old_pdf_path, metadata in valid_entries:
            new_filename = file_name_from_metadata(metadata, metadata.hash_file)
            new_pdf_path = processed_path / new_filename
            new_metadata_path = new_pdf_path.with_suffix(".json")

            if old_pdf_path == new_pdf_path:
                continue

            try:
                old_metadata_path = old_pdf_path.with_suffix(".json")
                shutil.move(old_pdf_path, new_pdf_path)
                shutil.move(old_metadata_path, new_metadata_path)
                renamed_count += 1
                logger.debug(f"Renamed: {old_pdf_path.name} -> {new_filename}")
            except Exception as e:
                logger.error(f"Failed to rename {old_pdf_path.name}: {e}")

        if not quiet:
            console.success(f"{len(valid_entries)} files validated, {renamed_count} renamed", indent=False)
        logger.debug(f"Renaming complete. Renamed {renamed_count} files.")

    return {'validated': len(valid_entries), 'renamed': renamed_count, 'orphans': orphan_count}


# --- Export prefix rules ---

def _get_nested_value(metadata: dict, key: str):
    """Get a value from metadata using dot notation (e.g., 'qrcode.qr_type')."""
    parts = key.split(".")
    current = metadata
    for part in parts:
        if not isinstance(current, dict) or part not in current:
            return None
        current = current[part]
    return current


def _match_value(actual, pattern: str) -> bool:
    """Match a metadata value against a pattern (supports trailing wildcard and numeric operators)."""
    if actual is None:
        return False
    for op in ('>=', '<=', '!=', '>', '<'):
        if pattern.startswith(op):
            try:
                return {
                    '>':  float(actual) > float(pattern[len(op):]),
                    '<':  float(actual) < float(pattern[len(op):]),
                    '>=': float(actual) >= float(pattern[len(op):]),
                    '<=': float(actual) <= float(pattern[len(op):]),
                    '!=': float(actual) != float(pattern[len(op):]),
                }[op]
            except (ValueError, TypeError):
                return False
    actual_str = str(actual).lower()
    if pattern.endswith("*"):
        return actual_str.startswith(pattern.lower()[:-1])
    if isinstance(actual, (int, float)):
        try:
            return float(actual) == float(pattern)
        except (ValueError, TypeError):
            pass
    return actual_str == pattern.lower()


def _resolve_match_value(pattern: str, profile_context: Optional[dict]) -> str:
    """Resolve ${profile.*} variables in a match pattern."""
    if not profile_context or not pattern.startswith("${profile."):
        return pattern
    key = pattern[len("${profile."):-1]
    value = profile_context.get(key)
    return value if value is not None else pattern


def _evaluate_export_prefix(metadata: dict, config, profile_context: Optional[dict] = None) -> str:
    """Evaluate export prefix rules against metadata. First match wins."""
    for rule in config.rules:
        if all(
            _match_value(_get_nested_value(metadata, k), _resolve_match_value(v, profile_context))
            for k, v in rule.match.items()
        ):
            return rule.prefix
    return config.default_prefix


def _build_filename_from_fields(metadata: dict, fields: list, file_hash: str) -> str:
    """Build a filename from selected metadata fields."""
    parts = []
    for field_name in fields:
        value = _get_nested_value(metadata, field_name)
        if value is not None and str(value).strip():
            component = sanitize_filename_component(str(value))
            if len(component) > 80:
                component = component[:80].rsplit(" ", 1)[0]
            parts.append(component)
    ext = metadata.get("source_extension") or ".pdf"
    parts.append(f"{file_hash}{ext}")
    return " - ".join(parts).lower()


def _should_skip_copy(src: Path, dst: Path) -> bool:
    """Check if a copy can be skipped (destination exists with same content)."""
    return (dst.exists()
            and src.stat().st_size == dst.stat().st_size
            and hash_file_fast(src) == hash_file_fast(dst))


def _check_file_size(src: Path, max_file_size_mb: float | None) -> None:
    """Log a warning if file exceeds the configured max size threshold."""
    if max_file_size_mb is None:
        return
    size = src.stat().st_size
    threshold = max_file_size_mb * 1024 * 1024
    if size >= threshold:
        size_mb = size / (1024 * 1024)
        logger.warning(f"Large file: {src.name} ({size_mb:.1f} MB exceeds {max_file_size_mb} MB threshold)")


# --- Copy / export ---

def copy_matching_files(
    processed_path: Path,
    pattern: str,
    dest_folder: Path,
    incremental: bool = False,
    export_config=None,
    profile_context: Optional[dict] = None,
    quiet: bool = False,
) -> dict:
    """Copy files matching pattern to destination."""
    from papertrail.utils import make_matcher

    console = get_console()
    dest_folder.mkdir(parents=True, exist_ok=True)
    matcher = make_matcher(pattern, use_search=True)
    stats = {'copied': 0, 'skipped': 0, 'deduped': 0, 'total': 0}
    seen_content_hashes = set()

    file_mappings = export_config.file_mappings if export_config is not None else None
    max_file_size_mb = export_config.max_file_size_mb if export_config is not None else None
    use_prefixes = file_mappings is not None and file_mappings.enabled

    if use_prefixes:
        matching_pdfs = []
        for file in processed_path.iterdir():
            if not file.is_file():
                continue
            if file.suffix.lower() not in (".pdf", ".xlsx"):
                continue
            if not matcher(file.name):
                continue
            matching_pdfs.append(file)

        for pdf_file in console.track(matching_pdfs, "Copying files"):
            stats['total'] += 1
            json_file = pdf_file.with_suffix(".json")

            metadata = {}
            if json_file.exists():
                metadata = load_json_data(json_file)

            ch = metadata.get("hash_content")
            if ch and ch in seen_content_hashes:
                stats['deduped'] += 1
                logger.debug(f"[EXPORT-DEDUP] Skipping {pdf_file.name} (content hash {ch} already exported)")
                continue
            if ch:
                seen_content_hashes.add(ch)

            prefix = _evaluate_export_prefix(metadata, file_mappings, profile_context)

            if file_mappings.filename_fields and metadata:
                file_hash = metadata.get("hash_file", pdf_file.stem.split(" - ")[-1])
                base_name = _build_filename_from_fields(
                    metadata, file_mappings.filename_fields, file_hash
                )
            else:
                base_name = pdf_file.name

            dest_pdf_name = prefix + base_name
            dest_json_name = Path(dest_pdf_name).with_suffix(".json").name

            dest_pdf = dest_folder / dest_pdf_name
            dest_json = dest_folder / dest_json_name

            if incremental and _should_skip_copy(pdf_file, dest_pdf):
                stats['skipped'] += 1
                continue

            _check_file_size(pdf_file, max_file_size_mb)
            shutil.copy2(pdf_file, dest_pdf)
            if json_file.exists():
                metadata['source_filename'] = pdf_file.name
                save_json_data(dest_json, metadata)
            stats['copied'] += 1

        if not quiet:
            console.success(f"Copied {stats['copied']} files to {dest_folder.name}", indent=False)
    else:
        matching_files = []
        for file in processed_path.iterdir():
            if not file.is_file():
                continue
            if file.suffix.lower() not in [".pdf", ".xlsx", ".json"]:
                continue
            if not matcher(file.name):
                continue
            matching_files.append(file)

        deduped_stems = set()
        for file in console.track(matching_files, "Copying files"):
            stats['total'] += 1
            dest_file = dest_folder / file.name

            if file.suffix.lower() in (".pdf", ".xlsx"):
                json_sidecar = file.with_suffix(".json")
                if json_sidecar.exists():
                    sidecar_data = load_json_data(json_sidecar)
                    ch = sidecar_data.get("hash_content")
                    if ch and ch in seen_content_hashes:
                        stats['deduped'] += 1
                        deduped_stems.add(file.stem)
                        logger.debug(f"[EXPORT-DEDUP] Skipping {file.name} (content hash {ch} already exported)")
                        continue
                    if ch:
                        seen_content_hashes.add(ch)
            elif file.suffix.lower() == ".json" and file.stem in deduped_stems:
                stats['total'] -= 1
                continue

            if incremental and _should_skip_copy(file, dest_file):
                stats['skipped'] += 1
                continue

            if file.suffix.lower() == '.json':
                data = load_json_data(file)
                src_ext = data.get("source_extension") or ".pdf"
                data['source_filename'] = file.with_suffix(src_ext).name
                save_json_data(dest_file, data)
            else:
                _check_file_size(file, max_file_size_mb)
                shutil.copy2(file, dest_file)
            stats['copied'] += 1

        pdf_copied = stats['copied'] // 2 if stats['copied'] > 0 else 0
        if not quiet:
            console.success(f"Copied {pdf_copied} files to {dest_folder.name}", indent=False)

    return stats


def export_metadata_to_excel(processed_path: Path, excel_output_path: str, quiet: bool = False) -> dict:
    """Export metadata to an Excel file."""
    from papertrail.metadata import load_json_files_parallel

    console = get_console()
    metadata_list = []

    for metadata_path, metadata in load_json_files_parallel(processed_path, validate=True, show_progress=not quiet, progress_desc="Collecting metadata"):
        metadata_dict = metadata.model_dump()
        metadata_dict.pop("class_reasoning", None)

        pdf_path = metadata_path.with_suffix(".pdf")
        filename = pdf_path.name if pdf_path.exists() else ""
        metadata_dict["filename"] = filename
        metadata_dict["filename_length"] = len(filename)

        try:
            date_parts = metadata.date_issued.split('-')
            metadata_dict["year"] = int(date_parts[0])
            metadata_dict["month"] = int(date_parts[1])
        except (IndexError, ValueError, AttributeError):
            metadata_dict["year"] = None
            metadata_dict["month"] = None

        if isinstance(metadata_dict.get("document_type"), str):
            metadata_dict["document_type"] = clean_enum_string(metadata_dict["document_type"], "DocumentType")

        metadata_list.append(metadata_dict)

    if not metadata_list:
        if not quiet:
            console.warning("No valid metadata found to export", indent=False)
        return {'exported': 0}

    df = pd.DataFrame(metadata_list)
    ordered_cols = [
        "class_confidence", "date_issued", "year", "month", "hash_content", "hash_file",
        "filename", "filename_length", "page_count", "document_type", "document_type_raw",
        "document_title", "issuing_party", "issuing_party_raw",
        "total_amount", "total_amount_currency"
    ]
    extra_cols = [col for col in df.columns if col not in ordered_cols]
    df = df[ordered_cols + extra_cols]

    if "date_issued" in df.columns:
        df = df.sort_values(by="date_issued", ascending=False)

    with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        worksheet.freeze_panes = 'A2'

        from openpyxl.utils import get_column_letter
        for col in ordered_cols:
            if col in df.columns:
                col_idx = df.columns.get_loc(col) + 1
                col_letter = get_column_letter(col_idx)
                values_lens = [len(str(val)) for val in df[col].values if val is not None]
                max_len = max(values_lens + [len(col)])
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 102)

        hidden_cols = ["year", "month", "filename_length"]
        for col in hidden_cols:
            if col in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                worksheet.column_dimensions[col_letter].hidden = True

    if not quiet:
        console.success(f"Exported {len(df)} entries", indent=False)
    logger.debug(f"Exported {len(df)} entries to {excel_output_path}")
    return {'exported': len(df)}


def _calculate_directory_hash(directory: Path) -> str:
    """Calculate a hash representing all PDF files in the directory."""
    pdf_files = sorted(directory.glob("*.pdf"))
    if not pdf_files:
        return ""
    combined = [f"{f.name}:{hash_file_fast(f)}" for f in pdf_files]
    return hashlib.sha256("\n".join(combined).encode()).hexdigest()[:16]


def _directory_has_changed(directory: Path) -> bool:
    """Check if directory contents have changed since last check."""
    hash_file_path = directory / ".directory_hash"
    current_hash = _calculate_directory_hash(directory)

    if not current_hash:
        return False

    if not hash_file_path.exists():
        with open(hash_file_path, "w") as f:
            f.write(current_hash)
        return True

    with open(hash_file_path, "r") as f:
        stored_hash = f.read().strip()

    if current_hash != stored_hash:
        with open(hash_file_path, "w") as f:
            f.write(current_hash)
        return True

    return False


def task_export_all_dates(
    processed_path: Path,
    export_base_dir: Path,
    run_merge: bool = False,
    export_config=None,
    profile_context: dict | None = None,
):
    """Export files for all unique dates found in processed files."""
    from papertrail.tasks.check import validate_merged_pdf

    console = get_console()
    processed_path = Path(processed_path)
    export_base_dir = Path(export_base_dir)

    setup_task_logging(processed_path, "export_all_dates")
    logger.debug("Scanning for unique dates in processed files...")
    all_dates = get_unique_dates(processed_path)

    if not all_dates:
        console.warning("No dates found in processed files", indent=False)
        return

    logger.debug(f"Found {len(all_dates)} unique dates: {', '.join(all_dates[:10])}{' ...' if len(all_dates) > 10 else ''}")

    total_copied = 0
    total_skipped = 0
    changed_directories = []

    for date in console.track(all_dates, "Exporting dates"):
        export_date_dir = export_base_dir / date
        logger.debug(f"[{date}] Processing...")

        if export_date_dir.exists():
            shutil.rmtree(export_date_dir)

        stats = copy_matching_files(processed_path, date, export_date_dir, incremental=False, export_config=export_config, profile_context=profile_context)
        total_copied += stats['copied']
        total_skipped += stats['skipped']

        if stats['total'] == 0:
            logger.debug(f"No files match date pattern '{date}'")
        else:
            logger.debug(f"Copied: {stats['copied']}, Skipped: {stats['skipped']}, Total: {stats['total']}")

        if stats['copied'] > 0:
            changed_directories.append(export_date_dir)
        elif stats['total'] > 0:
            if export_date_dir.exists() and _directory_has_changed(export_date_dir):
                changed_directories.append(export_date_dir)

    console.success(f"{len(all_dates)} dates exported, {total_copied} files copied", indent=False)
    logger.debug(f"Processed {len(all_dates)} date(s), Total files copied: {total_copied}, Skipped: {total_skipped}")

    if run_merge and changed_directories:
        logger.debug("=== Running PDF Merge ===")
        from pdf_gluer import merge_all_pdfs

        for export_dir in console.track(changed_directories, "Merging PDFs"):
            logger.debug(f"Merging PDFs in {export_dir}...")
            try:
                merge_all_pdfs(str(export_dir))
                logger.debug("Merge completed successfully")
                validate_merged_pdf(export_dir)
            except Exception as e:
                logger.error(f"Merge failed: {e}")

        console.success(f"Merged {len(changed_directories)} directories", indent=False)

    logger.debug("Export all dates complete.")


# --- Archive ---

def task_archive(processed_path: Path, digests: list[str], dry_run: bool = False) -> None:
    """Archive documents by hash_file digest."""
    console = get_console()
    archive_dir = processed_path.parent / "_archived"

    hash_to_json: dict[str, Path] = {}
    for json_path, data in iter_json_files(processed_path):
        hf = data.get("hash_file")
        if hf:
            hash_to_json[hf] = json_path

    if not dry_run:
        archive_dir.mkdir(exist_ok=True)

    if dry_run:
        console.info("Dry run — no files will be moved", indent=False)

    found = 0
    moved = 0
    not_found = []

    for digest in digests:
        json_path = hash_to_json.get(digest)
        if not json_path:
            not_found.append(digest)
            console.warning(f"[NOT FOUND] {digest}", indent=False)
            continue

        found += 1
        data = None
        try:
            data = load_json_data(json_path)
        except Exception:
            pass

        files_to_move = [json_path]

        companion = find_companion_file(json_path, data)
        if companion and companion.exists():
            files_to_move.append(companion)

        stem = json_path.stem
        for extra_suffix in (".embeddings.json", ".reconciliation.json"):
            extra = json_path.parent / (stem + extra_suffix)
            if extra.exists():
                files_to_move.append(extra)

        for src in files_to_move:
            dst = archive_dir / src.name
            if dst.exists():
                base, suffix = dst.stem, dst.suffix
                counter = 2
                while dst.exists():
                    dst = archive_dir / f"{base}_{counter}{suffix}"
                    counter += 1

            if dry_run:
                console.detail(f"[WOULD MOVE] {src.name}", indent=False)
            else:
                src.rename(dst)
                console.detail(f"[MOVED] {src.name}", indent=False)

        moved += 1

    console.info(
        f"Archive: {found} found, {moved} archived, {len(not_found)} not found",
        indent=False,
    )
    if not dry_run and moved > 0:
        console.detail(f"Archived to: {archive_dir}", indent=False)


# --- Gmail download ---

def task_gmail_download(months: int = 2):
    """Download email attachments from Gmail."""
    from papertrail.gmail import download_gmail_attachments
    from papertrail.utils import compute_month_range, month_to_date_range

    console = get_console()

    profile = get_current_profile()
    if not profile:
        console.error("No profile is active.", indent=False)
        raise RuntimeError("No profile is active.")

    raw_paths = profile.paths.raw
    processed_path_str = profile.paths.processed

    if processed_path_str:
        setup_task_logging(Path(processed_path_str), "gmail_download")

    if not raw_paths or not processed_path_str:
        missing = []
        if not raw_paths:
            missing.append("paths.raw")
        if not processed_path_str:
            missing.append("paths.processed")
        console.error(f"Missing required profile settings: {', '.join(missing)}", indent=False)
        raise RuntimeError(f"Missing required profile settings: {', '.join(missing)}")

    raw_path = Path(raw_paths[0])
    processed_path = Path(processed_path_str)

    export_dates = compute_month_range(months)

    totals = {
        "messages_found": 0, "messages_processed": 0, "messages_skipped": 0,
        "attachments_downloaded": 0, "attachments_failed": 0, "bytes_downloaded": 0,
    }

    gmail_dir = raw_path / "gmail"

    for month in export_dates:
        month_dir = gmail_dir / month
        month_dir.mkdir(parents=True, exist_ok=True)

        start_date, end_date = month_to_date_range([month])
        logger.debug(f"Gmail {month}: {start_date.date()} to {end_date.date()} → {month_dir}")

        try:
            stats = download_gmail_attachments(
                output_dir=month_dir,
                start_date=start_date,
                end_date=end_date,
                tracking_dir=gmail_dir,
            )
        except FileNotFoundError as e:
            console.error(f"Gmail credentials not found: {e}", indent=False)
            raise RuntimeError(f"Gmail credentials not found: {e}") from e
        except Exception as e:
            error_type = type(e).__name__
            console.error(f"Gmail download failed ({error_type}): {e}", indent=False)
            raise RuntimeError(f"Gmail download failed ({error_type}): {e}") from e

        for key in totals:
            totals[key] += stats[key]

    logger.debug(f"Messages found: {totals['messages_found']}")
    logger.debug(f"Messages processed: {totals['messages_processed']}")
    logger.debug(f"Messages skipped: {totals['messages_skipped']}")
    logger.debug(f"Attachments downloaded: {totals['attachments_downloaded']}")
    logger.debug(f"Attachments failed: {totals['attachments_failed']}")
    logger.debug(f"Bytes downloaded: {totals['bytes_downloaded']:,}")

    if totals['attachments_downloaded'] > 0:
        console.success(
            f"{totals['messages_processed']} messages processed, "
            f"{totals['attachments_downloaded']} new attachments",
            indent=False
        )
    elif totals['messages_processed'] > 0:
        console.success(
            f"{totals['messages_processed']} messages processed, 0 new attachments",
            indent=False
        )
    else:
        date_range = f"{export_dates[0]} to {export_dates[-1]}"
        console.warning(f"No messages found ({date_range})", indent=False)


# --- Merge attachments ---

def _match_type_pattern(doc_type: str, pattern: str) -> bool:
    """Check if doc_type matches a pattern with pipe-separated alternatives and trailing * wildcard."""
    doc_lower = doc_type.lower()
    for alt in pattern.split("|"):
        alt = alt.strip().lower()
        if alt.endswith("*"):
            if doc_lower.startswith(alt[:-1]):
                return True
        elif doc_lower == alt:
            return True
    return False


def merge_reconciled_attachments(
    export_path: Path,
    all_matches: list,
    merge_rules: list,
) -> dict:
    """Merge attachment PDFs into target PDFs based on reconciliation match data."""
    stats = {"merged": 0, "skipped": 0, "errors": 0}

    if not merge_rules or not all_matches:
        return stats

    merged_attachments: set[str] = set()

    for match in all_matches:
        candidates = match.pdf_candidates

        for rule in merge_rules:
            targets = [
                c for c in candidates
                if c.document_type and _match_type_pattern(c.document_type, rule.target_type)
                and not c.is_sub_document
            ]
            attachments = [
                c for c in candidates
                if c.document_type and _match_type_pattern(c.document_type, rule.attach_type)
                and not c.is_sub_document
            ]

            if not targets or not attachments:
                continue

            for target in targets:
                target_pdf = export_path / target.pdf_filename
                if not target_pdf.exists() or target_pdf.suffix.lower() != ".pdf":
                    continue

                for attachment in attachments:
                    if attachment.pdf_filename in merged_attachments:
                        logger.debug(
                            f"[MERGE] Skipping {attachment.pdf_filename} — already merged elsewhere"
                        )
                        stats["skipped"] += 1
                        continue

                    attach_pdf = export_path / attachment.pdf_filename
                    if not attach_pdf.exists() or attach_pdf.suffix.lower() != ".pdf":
                        continue

                    if target_pdf == attach_pdf:
                        continue

                    try:
                        import pikepdf

                        with pikepdf.open(target_pdf, allow_overwriting_input=True) as target_doc:
                            with pikepdf.open(attach_pdf) as attach_doc:
                                target_doc.pages.extend(attach_doc.pages)
                            target_doc.save(target_pdf)

                        merged_attachments.add(attachment.pdf_filename)
                        stats["merged"] += 1
                        logger.debug(
                            f"[MERGE] Appended {attachment.pdf_filename} "
                            f"to {target.pdf_filename}"
                        )
                    except Exception as e:
                        stats["errors"] += 1
                        logger.error(
                            f"[MERGE] Failed to append {attachment.pdf_filename} "
                            f"to {target.pdf_filename}: {e}"
                        )

    return stats
