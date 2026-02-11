"""Export tasks."""

import hashlib
import os
import shutil
from pathlib import Path

import pandas as pd

from papertrail.console import get_console
from papertrail.hashing import hash_file_fast
from papertrail.logging_utils import get_logger, setup_task_logging
from papertrail.metadata import get_unique_dates
from papertrail.models import normalize_enum_field_in_dict
from papertrail.tasks.organization import copy_matching_files
from papertrail.tasks.validation import validate_merged_pdf

logger = get_logger('cli')


def export_metadata_to_excel(processed_path: Path, excel_output_path: str, quiet: bool = False) -> dict:
    """Export metadata to an Excel file."""
    from papertrail.metadata import load_json_files_parallel

    console = get_console()
    metadata_list = []

    for metadata_path, metadata in load_json_files_parallel(processed_path, validate=True, show_progress=not quiet, progress_desc="Collecting metadata"):
        metadata_dict = metadata.model_dump()

        metadata_dict.pop("class_reasoning", None)

        pdf_path = metadata_path.with_suffix(".pdf")
        filename = pdf_path.name if pdf_path.exists() else ""
        metadata_dict["filename"] = filename
        metadata_dict["filename_length"] = len(filename)

        try:
            date_parts = metadata.date_issued.split('-')
            metadata_dict["year"] = int(date_parts[0])
            metadata_dict["month"] = int(date_parts[1])
        except (IndexError, ValueError, AttributeError):
            metadata_dict["year"] = None
            metadata_dict["month"] = None

        normalize_enum_field_in_dict(metadata_dict, "document_type", "DocumentType")

        metadata_list.append(metadata_dict)

    if metadata_list:
        df = pd.DataFrame(metadata_list)
        ordered_cols = [
            "class_confidence", "date_issued", "year", "month", "hash_content", "hash_file",
            "filename", "filename_length", "page_count", "document_type", "document_type_raw",
            "document_title", "issuing_party", "issuing_party_raw",
            "total_amount", "total_amount_currency"
        ]
        extra_cols = [col for col in df.columns if col not in ordered_cols]
        df = df[ordered_cols + extra_cols]

        if "date_issued" in df.columns:
            df = df.sort_values(by="date_issued", ascending=False)

        with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            worksheet.freeze_panes = 'A2'

            from openpyxl.utils import get_column_letter
            for col in ordered_cols:
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1
                    col_letter = get_column_letter(col_idx)
                    values_lens = [len(str(val)) for val in df[col].values if val is not None]
                    max_len = max(values_lens + [len(col)])
                    worksheet.column_dimensions[col_letter].width = min(max_len + 2, 102)

            hidden_cols = ["year", "month", "filename_length"]
            for col in hidden_cols:
                if col in df.columns:
                    col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                    worksheet.column_dimensions[col_letter].hidden = True

        if not quiet:
            console.success(f"Exported {len(df)} entries", indent=False)
        logger.debug(f"Exported {len(df)} entries to {excel_output_path}")
        return {'exported': len(df)}
    else:
        if not quiet:
            console.warning("No valid metadata found to export", indent=False)
        return {'exported': 0}


def calculate_directory_hash(directory: Path) -> str:
    """Calculate a hash representing all PDF files in the directory."""
    pdf_files = sorted(directory.glob("*.pdf"))
    if not pdf_files:
        return ""

    combined = []
    for pdf_file in pdf_files:
        file_hash = hash_file_fast(pdf_file)
        combined.append(f"{pdf_file.name}:{file_hash}")

    combined_str = "\n".join(combined)
    return hashlib.sha256(combined_str.encode()).hexdigest()[:16]


def directory_has_changed(directory: Path) -> bool:
    """Check if directory contents have changed since last check."""
    hash_file_path = directory / ".directory_hash"
    current_hash = calculate_directory_hash(directory)

    if not current_hash:
        return False

    if not hash_file_path.exists():
        with open(hash_file_path, "w") as f:
            f.write(current_hash)
        return True

    with open(hash_file_path, "r") as f:
        stored_hash = f.read().strip()

    if current_hash != stored_hash:
        with open(hash_file_path, "w") as f:
            f.write(current_hash)
        return True

    return False


def task_export_all_dates(
    processed_path: Path,
    export_base_dir: Path,
    run_merge: bool = False,
    export_config=None,
    profile_context: dict | None = None,
):
    """Export files for all unique dates found in processed files."""
    console = get_console()
    processed_path = Path(processed_path)
    export_base_dir = Path(export_base_dir)

    setup_task_logging(processed_path, "export_all_dates")
    logger.debug("Scanning for unique dates in processed files...")
    all_dates = get_unique_dates(processed_path)

    if not all_dates:
        console.warning("No dates found in processed files", indent=False)
        return

    logger.debug(f"Found {len(all_dates)} unique dates: {', '.join(all_dates[:10])}{' ...' if len(all_dates) > 10 else ''}")

    total_copied = 0
    total_skipped = 0
    changed_directories = []

    for date in console.track(all_dates, "Exporting dates"):
        export_date_dir = export_base_dir / date
        logger.debug(f"[{date}] Processing...")

        # Purge before export to avoid stale content
        if export_date_dir.exists():
            shutil.rmtree(export_date_dir)

        stats = copy_matching_files(processed_path, date, export_date_dir, incremental=False, export_config=export_config, profile_context=profile_context)
        total_copied += stats['copied']
        total_skipped += stats['skipped']

        if stats['total'] == 0:
            logger.debug(f"No files match date pattern '{date}'")
        else:
            logger.debug(f"Copied: {stats['copied']}, Skipped: {stats['skipped']}, Total: {stats['total']}")

        if stats['copied'] > 0:
            changed_directories.append(export_date_dir)
        elif stats['total'] > 0:
            if export_date_dir.exists() and directory_has_changed(export_date_dir):
                changed_directories.append(export_date_dir)

    # Summary
    console.success(f"{len(all_dates)} dates exported, {total_copied} files copied", indent=False)
    logger.debug(f"Processed {len(all_dates)} date(s), Total files copied: {total_copied}, Skipped: {total_skipped}")

    if run_merge and changed_directories:
        logger.debug("=== Running PDF Merge ===")
        from pdf_gluer import merge_all_pdfs

        for export_dir in console.track(changed_directories, "Merging PDFs"):
            logger.debug(f"Merging PDFs in {export_dir}...")
            try:
                merge_all_pdfs(str(export_dir))
                logger.debug("Merge completed successfully")
                validate_merged_pdf(export_dir)
            except Exception as e:
                logger.error(f"Merge failed: {e}")

        console.success(f"Merged {len(changed_directories)} directories", indent=False)

    logger.debug("Export all dates complete.")
