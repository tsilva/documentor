"""Canonical workflow entrypoints for the papertrail application."""

from __future__ import annotations

import fcntl
import hashlib
import io
import os
import re
import shutil
import sys
import time
import warnings
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Optional

import pandas as pd

from papertrail.app import App
from papertrail.console import get_console
from papertrail.hashing import hash_file_fast
from papertrail.logging_utils import get_logger, setup_failure_logger, setup_task_logging, suppress_console_logging
from papertrail.models import DocumentMetadata, clean_enum_string
from papertrail.naming import sanitize_filename_component
from papertrail.rules import RuleEngine
from papertrail.store import DocumentStore
from papertrail.workflow_utils import task_log_context

logger = get_logger("cli")


def extract(
    app: App,
    processed_path: Path,
    raw_paths: list[Path],
    quiet: bool = False,
) -> dict | None:
    """Extract and classify new PDF/XLSX/image files."""
    from papertrail.tasks.extraction import DocumentService

    lock_path = app.paths.cache / ".extract.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_file = open(lock_path, "w")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError:
        logger.error("Another extract process is already running. Exiting.")
        lock_file.close()
        return None

    try:
        with task_log_context(processed_path, "extract_new", show_header=not quiet):
            logs_dir = processed_path / "logs"
            failure_log_path = logs_dir / "classification_failures.log"
            failure_logger = setup_failure_logger(failure_log_path)
            logger.debug(f"Logging failures to: {failure_log_path}")
            return DocumentService(app).extract_new(
                processed_path,
                raw_paths,
                quiet=quiet,
                failure_logger=failure_logger,
            )
    finally:
        fcntl.flock(lock_file, fcntl.LOCK_UN)
        lock_file.close()


def sync(
    app: App,
    processed_path: Path,
    dry_run: bool = False,
    all_unknown: bool = False,
    pattern: str | None = None,
    workers: int = 1,
    all: bool = False,
    quiet: bool = False,
) -> dict:
    """Sync metadata by reclassifying in place."""
    from papertrail.tasks.extraction import DocumentService

    with task_log_context(processed_path, "sync", show_header=not quiet):
        return DocumentService(app).sync(
            processed_path,
            dry_run=dry_run,
            all_unknown=all_unknown,
            pattern=pattern,
            workers=workers,
            all=all,
            quiet=quiet,
        )


def rename(app: App, processed_path: Path, quiet: bool = False) -> dict:
    """Rename existing document files based on metadata."""
    store = DocumentStore(app)
    console = app.console

    with task_log_context(processed_path, "rename_files", show_header=not quiet):
        stats = store.repair_filenames(processed_path)
        if not quiet:
            console.success(
                f"{stats['validated']} files validated, {stats['renamed']} renamed",
                indent=False,
            )
        logger.debug(f"Renaming complete. Renamed {stats['renamed']} files.")
        return stats


def _build_filename_from_fields(metadata: dict, fields: list[str], file_hash: str) -> str:
    engine = RuleEngine()
    parts = []
    for field_name in fields:
        value = engine.get_nested_value(metadata, field_name)
        if value is not None and str(value).strip():
            component = engine.resolve_profile_value(str(value))
            component = sanitize_filename_component(component.strip())
            if len(component) > 80:
                component = component[:80].rsplit(" ", 1)[0]
            parts.append(component)
    ext = metadata.get("source_extension") or ".pdf"
    parts.append(f"{file_hash}{ext}")
    return " - ".join(parts).lower()


def _should_skip_copy(src: Path, dst: Path) -> bool:
    return dst.exists() and src.stat().st_size == dst.stat().st_size and hash_file_fast(src) == hash_file_fast(dst)


def _check_file_size(src: Path, max_file_size_mb: float | None) -> None:
    if max_file_size_mb is None:
        return
    size = src.stat().st_size
    threshold = max_file_size_mb * 1024 * 1024
    if size >= threshold:
        size_mb = size / (1024 * 1024)
        logger.warning(f"Large file: {src.name} ({size_mb:.1f} MB exceeds {max_file_size_mb} MB threshold)")


def copy_matching(
    app: App,
    processed_path: Path,
    pattern: str,
    dest_folder: Path,
    incremental: bool = False,
    export_config=None,
    profile_context: Optional[dict] = None,
    quiet: bool = False,
) -> dict:
    """Copy matching documents and their sidecars to a destination."""
    from papertrail.utils import make_matcher

    store = DocumentStore(app)
    console = app.console
    matcher = make_matcher(pattern, use_search=True)
    dest_folder.mkdir(parents=True, exist_ok=True)

    file_mappings = export_config.file_mappings if export_config is not None else None
    max_file_size_mb = export_config.max_file_size_mb if export_config is not None else None
    use_prefixes = file_mappings is not None and file_mappings.enabled
    engine = RuleEngine(profile_context=profile_context)

    stats = {"copied": 0, "skipped": 0, "deduped": 0, "total": 0}
    seen_content_hashes: set[str] = set()

    documents = list(store.iter_documents(processed_path, validate=False, require_companion=True))
    for json_path, doc_path, metadata in console.track(documents, "Copying files"):
        metadata_dict = metadata.model_dump() if isinstance(metadata, DocumentMetadata) else metadata
        if not matcher(doc_path.name) and not matcher(json_path.name):
            continue

        stats["total"] += 1
        content_hash = metadata_dict.get("hash_content")
        if content_hash and content_hash in seen_content_hashes:
            stats["deduped"] += 1
            logger.debug(f"[EXPORT-DEDUP] Skipping {doc_path.name} (content hash {content_hash} already exported)")
            continue
        if content_hash:
            seen_content_hashes.add(content_hash)

        if use_prefixes:
            prefix = engine.evaluate_export_prefix(metadata_dict, file_mappings=file_mappings)
            if file_mappings.filename_fields:
                file_hash = metadata_dict.get("hash_file", doc_path.stem.split(" - ")[-1])
                base_name = _build_filename_from_fields(metadata_dict, list(file_mappings.filename_fields), file_hash)
            else:
                base_name = doc_path.name
            dest_doc = dest_folder / f"{prefix}{base_name}"
            dest_json = dest_doc.with_suffix(".json")
        else:
            dest_doc = dest_folder / doc_path.name
            dest_json = dest_folder / json_path.name

        if incremental and _should_skip_copy(doc_path, dest_doc):
            stats["skipped"] += 1
            continue

        _check_file_size(doc_path, max_file_size_mb)
        shutil.copy2(doc_path, dest_doc)
        metadata_copy = dict(metadata_dict)
        metadata_copy["source_filename"] = doc_path.name
        store.save_json(dest_json, metadata_copy)
        stats["copied"] += 1

    if not quiet:
        console.success(f"Copied {stats['copied']} files to {dest_folder.name}", indent=False)
    return stats


def export_excel(app: App, processed_path: Path, excel_output_path: str, quiet: bool = False) -> dict:
    """Export metadata to an Excel file."""
    store = DocumentStore(app)
    console = app.console
    metadata_list = []

    for metadata_path, metadata in store.load_sidecars_parallel(
        processed_path,
        validate=True,
        show_progress=not quiet,
        progress_desc="Collecting metadata",
    ):
        metadata_dict = metadata.model_dump()
        metadata_dict.pop("class_reasoning", None)

        doc_path = store.find_companion(metadata_path, metadata_dict)
        filename = doc_path.name if doc_path and doc_path.exists() else ""
        metadata_dict["filename"] = filename
        metadata_dict["filename_length"] = len(filename)

        try:
            date_parts = metadata.date_issued.split("-")
            metadata_dict["year"] = int(date_parts[0])
            metadata_dict["month"] = int(date_parts[1])
        except (IndexError, ValueError, AttributeError):
            metadata_dict["year"] = None
            metadata_dict["month"] = None

        if isinstance(metadata_dict.get("document_type"), str):
            metadata_dict["document_type"] = clean_enum_string(metadata_dict["document_type"], "DocumentType")

        metadata_list.append(metadata_dict)

    if not metadata_list:
        if not quiet:
            console.warning("No valid metadata found to export", indent=False)
        return {"exported": 0}

    df = pd.DataFrame(metadata_list)
    ordered_cols = [
        "class_confidence", "date_issued", "year", "month", "hash_content", "hash_file",
        "filename", "filename_length", "page_count", "document_type", "document_type_raw",
        "document_title", "issuing_party", "issuing_party_raw",
        "total_amount", "total_amount_currency",
    ]
    extra_cols = [col for col in df.columns if col not in ordered_cols]
    df = df[ordered_cols + extra_cols]

    if "date_issued" in df.columns:
        df = df.sort_values(by="date_issued", ascending=False)

    with pd.ExcelWriter(excel_output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]
        worksheet.freeze_panes = "A2"

        from openpyxl.utils import get_column_letter

        for col in ordered_cols:
            if col in df.columns:
                col_idx = df.columns.get_loc(col) + 1
                col_letter = get_column_letter(col_idx)
                values_lens = [len(str(val)) for val in df[col].values if val is not None]
                max_len = max(values_lens + [len(col)])
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 102)

        for col in ("year", "month", "filename_length"):
            if col in df.columns:
                col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                worksheet.column_dimensions[col_letter].hidden = True

    if not quiet:
        console.success(f"Exported {len(df)} entries", indent=False)
    logger.debug(f"Exported {len(df)} entries to {excel_output_path}")
    return {"exported": len(df)}


def _calculate_directory_hash(directory: Path) -> str:
    pdf_files = sorted(directory.glob("*.pdf"))
    if not pdf_files:
        return ""
    combined = [f"{f.name}:{hash_file_fast(f)}" for f in pdf_files]
    return hashlib.sha256("\n".join(combined).encode()).hexdigest()[:16]


def _directory_has_changed(directory: Path) -> bool:
    hash_file_path = directory / ".directory_hash"
    current_hash = _calculate_directory_hash(directory)
    if not current_hash:
        return False
    if not hash_file_path.exists():
        hash_file_path.write_text(current_hash, encoding="utf-8")
        return True
    stored_hash = hash_file_path.read_text(encoding="utf-8").strip()
    if current_hash != stored_hash:
        hash_file_path.write_text(current_hash, encoding="utf-8")
        return True
    return False


def export_dates(
    app: App,
    processed_path: Path,
    export_base_dir: Path,
    run_merge: bool = False,
    export_config=None,
    profile_context: dict | None = None,
) -> None:
    """Export files for all unique dates found in processed files."""
    from papertrail.tasks.check import validate_merged_pdf

    store = DocumentStore(app)
    console = app.console

    with task_log_context(processed_path, "export_all_dates", show_header=False):
        all_dates = store.unique_dates(processed_path)
        if not all_dates:
            console.warning("No dates found in processed files", indent=False)
            return

        total_copied = 0
        total_skipped = 0
        changed_directories = []
        for date in console.track(all_dates, "Exporting dates"):
            export_date_dir = export_base_dir / date
            logger.debug(f"[{date}] Processing...")

            if export_date_dir.exists():
                shutil.rmtree(export_date_dir)

            stats = copy_matching(
                app,
                processed_path,
                date,
                export_date_dir,
                incremental=False,
                export_config=export_config,
                profile_context=profile_context,
                quiet=True,
            )
            total_copied += stats["copied"]
            total_skipped += stats["skipped"]

            if stats["copied"] > 0:
                changed_directories.append(export_date_dir)
            elif stats["total"] > 0 and export_date_dir.exists() and _directory_has_changed(export_date_dir):
                changed_directories.append(export_date_dir)

        console.success(f"{len(all_dates)} dates exported, {total_copied} files copied", indent=False)
        logger.debug(
            f"Processed {len(all_dates)} date(s), Total files copied: {total_copied}, Skipped: {total_skipped}"
        )

        if run_merge and changed_directories:
            from pdf_gluer import merge_all_pdfs

            logger.debug("=== Running PDF Merge ===")
            for export_dir in console.track(changed_directories, "Merging PDFs"):
                logger.debug(f"Merging PDFs in {export_dir}...")
                try:
                    merge_all_pdfs(str(export_dir))
                    validate_merged_pdf(export_dir)
                except Exception as exc:
                    logger.error(f"Merge failed: {exc}")

            console.success(f"Merged {len(changed_directories)} directories", indent=False)


def archive(app: App, processed_path: Path, digests: list[str], dry_run: bool = False) -> None:
    """Archive documents by hash_file digest."""
    console = app.console
    store = DocumentStore(app)

    if dry_run:
        console.info("Dry run - no files will be moved", indent=False)

    stats = store.archive_by_hash_file(digests, dry_run=dry_run, scope=processed_path)
    for digest in stats["not_found"]:
        console.warning(f"[NOT FOUND] {digest}", indent=False)

    console.info(
        f"Archive: {stats['found']} found, {stats['archived']} archived, {len(stats['not_found'])} not found",
        indent=False,
    )
    if not dry_run and stats["archived"] > 0:
        console.detail(f"Archived to: {stats['archive_dir']}", indent=False)


def gmail(app: App, months: int = 2) -> None:
    """Download email attachments from Gmail."""
    from papertrail.gmail import download_gmail_attachments
    from papertrail.utils import compute_month_range, month_to_date_range

    console = app.console
    raw_paths = app.profile.paths.raw
    processed_path_str = app.profile.paths.processed

    if processed_path_str:
        setup_task_logging(Path(processed_path_str), "gmail_download")

    if not raw_paths or not processed_path_str:
        missing = []
        if not raw_paths:
            missing.append("paths.raw")
        if not processed_path_str:
            missing.append("paths.processed")
        console.error(f"Missing required profile settings: {', '.join(missing)}", indent=False)
        raise RuntimeError(f"Missing required profile settings: {', '.join(missing)}")

    raw_path = Path(raw_paths[0])
    export_dates = compute_month_range(months)
    totals = {
        "messages_found": 0,
        "messages_processed": 0,
        "messages_skipped": 0,
        "attachments_downloaded": 0,
        "attachments_failed": 0,
        "bytes_downloaded": 0,
    }

    gmail_dir = raw_path / "gmail"
    for month in export_dates:
        month_dir = gmail_dir / month
        month_dir.mkdir(parents=True, exist_ok=True)
        start_date, end_date = month_to_date_range([month])
        logger.debug(f"Gmail {month}: {start_date.date()} to {end_date.date()} -> {month_dir}")

        try:
            stats = download_gmail_attachments(
                output_dir=month_dir,
                start_date=start_date,
                end_date=end_date,
                tracking_dir=gmail_dir,
            )
        except FileNotFoundError as exc:
            console.error(f"Gmail credentials not found: {exc}", indent=False)
            raise RuntimeError(f"Gmail credentials not found: {exc}") from exc
        except Exception as exc:
            error_type = type(exc).__name__
            console.error(f"Gmail download failed ({error_type}): {exc}", indent=False)
            raise RuntimeError(f"Gmail download failed ({error_type}): {exc}") from exc

        for key in totals:
            totals[key] += stats[key]

    if totals["attachments_downloaded"] > 0:
        console.success(
            f"{totals['messages_processed']} messages processed, {totals['attachments_downloaded']} new attachments",
            indent=False,
        )
    elif totals["messages_processed"] > 0:
        console.success(
            f"{totals['messages_processed']} messages processed, 0 new attachments",
            indent=False,
        )
    else:
        date_range = f"{export_dates[0]} to {export_dates[-1]}"
        console.warning(f"No messages found ({date_range})", indent=False)


def check(app: App, processed_path: Path, verify_hashes: bool = False, dry_run: bool = False) -> None:
    """Run the integrity and backfill checks."""
    from papertrail.tasks.check import task_check as _task_check_impl

    _task_check_impl(processed_path, verify_hashes=verify_hashes, dry_run=dry_run)


def validate_merged_pdf(folder_path: Path) -> bool:
    """Validate that a merged PDF has the expected page count."""
    from papertrail.tasks.check import validate_merged_pdf as _validate_merged_pdf

    return _validate_merged_pdf(folder_path)


def reconcile(
    app: App,
    export_path: Path,
    excel_path: Optional[Path] = None,
    dry_run: bool = False,
) -> None:
    """Reconcile bank transactions against exported documents."""
    from papertrail.tasks.reconciliation import task_reconcile as _task_reconcile_impl

    _task_reconcile_impl(export_path, excel_path=excel_path, dry_run=dry_run)


def pipeline(
    app: App,
    months: int = 2,
    export_date_arg: Optional[str] = None,
    processed_path_override: Optional[Path] = None,
) -> None:
    """Run the full document processing pipeline."""
    from papertrail.config import get_passwords
    from papertrail.mbox import extract_mbox_attachments
    from papertrail.tasks.check import validate_merged_pdf
    from papertrail.tasks.organize import merge_reconciled_attachments
    from papertrail.tasks.reconciliation import _discover_bank_statements, _reconcile_single
    from papertrail.utils import compute_month_range, month_to_date_range

    console = app.console
    start_time = time.time()
    profile = app.profile

    raw_dirs = profile.paths.raw
    processed_dir = str(processed_path_override or profile.paths.processed or "")
    export_dir = profile.paths.export

    missing = []
    if not raw_dirs:
        missing.append("paths.raw")
    if not processed_dir:
        missing.append("paths.processed")
    if not export_dir:
        missing.append("paths.export")
    if missing:
        console.error(f"Missing required profile settings: {', '.join(missing)}", indent=False)
        sys.exit(1)

    processed_path = Path(processed_dir)
    log_file_path = setup_task_logging(processed_path, "pipeline")
    console.pipeline_header(profile.profile.name, str(log_file_path))

    if export_date_arg:
        export_dates_list = [export_date_arg]
    else:
        export_dates_list = compute_month_range(months)

    for ed in export_dates_list:
        if not re.match(r"^\d{4}-\d{2}$", ed):
            console.error(f"The export_date must be in YYYY-MM format: {ed}", indent=False)
            sys.exit(1)

    passwords, _ = get_passwords()
    if not passwords:
        logger.debug("No passwords configured. Password-protected archives will be skipped.")

    processed_files_excel_path = processed_path / "processed_files.xlsx"
    pipeline_warnings: list[str] = []
    summary: dict[str, str] = {}
    output_paths: list[tuple[str, str]] = []

    if profile.gmail.enabled:
        with console.step_progress("Download Gmail attachments") as step:
            try:
                from papertrail.gmail import download_gmail_attachments

                raw_path = Path(raw_dirs[0])
                gmail_dir = raw_path / "gmail"
                totals = {
                    "messages_found": 0,
                    "messages_processed": 0,
                    "messages_skipped": 0,
                    "attachments_downloaded": 0,
                    "attachments_failed": 0,
                    "bytes_downloaded": 0,
                }
                for month in export_dates_list:
                    month_dir = gmail_dir / month
                    month_dir.mkdir(parents=True, exist_ok=True)
                    gmail_start, gmail_end = month_to_date_range([month])
                    month_stats = download_gmail_attachments(
                        output_dir=month_dir,
                        start_date=gmail_start,
                        end_date=gmail_end,
                        quiet=True,
                        tracking_dir=gmail_dir,
                    )
                    for key in totals:
                        totals[key] += month_stats[key]

                if totals["attachments_downloaded"] > 0:
                    step.success(f"{totals['messages_processed']} messages, {totals['attachments_downloaded']} new attachments")
                elif totals["messages_processed"] > 0:
                    step.success(f"{totals['messages_processed']} messages, 0 new attachments")
                else:
                    date_range = f"{export_dates_list[0]} to {export_dates_list[-1]}"
                    step.warning(f"No messages found ({date_range})")
            except Exception as exc:
                msg = f"Gmail download failed, continuing pipeline ({exc})"
                step.warning(msg)
                pipeline_warnings.append(msg)
                logger.warning(f"Gmail download failed (non-fatal): {exc}")

    for raw_dir in raw_dirs:
        with console.step_progress("Extract mbox attachments") as step:
            stats = extract_mbox_attachments(raw_dir)
            if stats["mbox_files"] > 0:
                step.success(f"{stats['mbox_files']} mbox file(s), {stats['attachments_extracted']} attachment(s)")
            else:
                step.warning("No mbox files found")
            if stats["errors"]:
                step.error(f"{len(stats['errors'])} error(s)")
                sys.exit(1)

        with console.step_progress("Extract compressed archives") as step:
            from archive_extractor import extract_archives

            with redirect_stdout(io.StringIO()), redirect_stderr(io.StringIO()), warnings.catch_warnings():
                warnings.simplefilter("ignore")
                results = extract_archives(raw_dir, passwords=passwords if passwords else None)
            total_extracted = 0
            failures = 0
            for _, count in results.items():
                if count == -1:
                    failures += 1
                else:
                    total_extracted += count
            if total_extracted > 0:
                step.success(f"{total_extracted} files from {len(results) - failures} archive(s)")
            elif failures > 0:
                step.warning(f"{failures} archive(s) failed")
            else:
                step.warning("No archives found")

    console.step("Classify new documents")
    try:
        extract_stats = extract(app, processed_path, [Path(d) for d in raw_dirs], quiet=False)
        if extract_stats is None:
            console.warning("Extraction locked by another process")
        elif extract_stats["failed"] > 0:
            console.error(f"OpenRouter API error ({extract_stats['failed']} failed)")
    except Exception as exc:
        console.error(str(exc))
        sys.exit(1)

    if extract_stats and extract_stats.get("bundles_split", 0) > 0:
        console.notable(f"Split {extract_stats['bundles_split']} PDF bundles into {extract_stats['split_pages']} pages")
    if extract_stats:
        if extract_stats.get("new_issuers"):
            console.notable(f"New issuers: {', '.join(extract_stats['new_issuers'])}")
        unknown_parts = []
        if extract_stats.get("unknown_document_type"):
            unknown_parts.append(f"{extract_stats['unknown_document_type']} document_type")
        if extract_stats.get("unknown_issuing_party"):
            unknown_parts.append(f"{extract_stats['unknown_issuing_party']} issuing_party")
        if unknown_parts:
            unknown_msg = f"Unknown: {', '.join(unknown_parts)}"
            console.notable(unknown_msg)
            pipeline_warnings.append(f"{unknown_msg} in extracted documents")

        if extract_stats["new"] > 0 or extract_stats["duplicates"] > 0:
            classified_parts = [f"{extract_stats['new']} new", f"{extract_stats['duplicates']} duplicates"]
            if extract_stats.get("batch_duplicates", 0) > 0:
                classified_parts.append(f"{extract_stats['batch_duplicates']} batch dupes")
            summary["Classified"] = ", ".join(classified_parts)

    console.step("Sync orphaned metadata")
    try:
        sync_stats = sync(app, processed_path, quiet=False)
        orphans = sync_stats.get("targets", 0)
        if orphans == 0:
            console.success("0 orphans found")
        else:
            resynced = sync_stats.get("new", 0) + sync_stats.get("changed", 0)
            console.success(f"{resynced} orphans re-synced")
    except Exception as exc:
        console.error(str(exc))
        sys.exit(1)

    with console.step_progress("Sync filenames to metadata") as step:
        try:
            rename_stats = rename(app, processed_path, quiet=True)
            step.success(f"{rename_stats['validated']} validated, {rename_stats['renamed']} renamed")
            if rename_stats["renamed"] > 0:
                summary["Renamed"] = f"{rename_stats['renamed']} files"
            orphans = rename_stats.get("orphans", 0)
            if orphans > 0:
                msg = f"{orphans} orphaned JSON sidecar(s) (companion file missing)"
                console.notable(msg)
                pipeline_warnings.append(msg)
        except Exception as exc:
            step.error(str(exc))
            sys.exit(1)

    with console.step_progress("Export to Excel") as step:
        try:
            excel_stats = export_excel(app, processed_path, str(processed_files_excel_path), quiet=True)
            if excel_stats["exported"]:
                step.success(f"{excel_stats['exported']} entries")
                summary["Exported"] = f"{excel_stats['exported']} entries"
            else:
                step.warning("No valid metadata found to export")
        except Exception as exc:
            step.error(str(exc))
            sys.exit(1)

    export_file_config = profile.export
    profile_context = {"tax_number": profile.profile.tax_number} if profile.profile.tax_number else None
    recon_stats_all: list[dict] = []
    merge_rules = export_file_config.merge_rules

    for export_date in export_dates_list:
        export_date_dir = os.path.join(export_dir, export_date)
        if os.path.exists(export_date_dir):
            shutil.rmtree(export_date_dir)

        with console.step_progress(f"Export documents ({export_date})") as step:
            try:
                copy_stats = copy_matching(
                    app,
                    processed_path,
                    export_date,
                    Path(export_date_dir),
                    export_config=export_file_config,
                    profile_context=profile_context,
                    quiet=True,
                )
                copied = copy_stats.get("copied", 0)
                deduped = copy_stats.get("deduped", 0)
                if copied:
                    msg = f"{copied} files"
                    if deduped > 0:
                        msg += f" ({deduped} content dupes skipped)"
                    step.success(msg)
                else:
                    step.success("0 files")
            except Exception as exc:
                step.error(str(exc))
                sys.exit(1)

        all_recon_matches = []
        bank_statements = _discover_bank_statements(Path(export_date_dir))
        for statement_path in bank_statements:
            statement_json = statement_path.with_suffix(".json")
            statement_info = {}
            if statement_json.exists():
                try:
                    statement_info = DocumentStore(app).load_metadata(statement_json).get("bank_statement", {}) or {}
                except Exception:
                    pass
            account = statement_info.get("account_number", statement_path.stem)
            period = statement_info.get("period_start", export_date)
            if period and len(period) >= 7:
                period = period[:7]

            with console.step_progress(f"Match bank transactions: {account} ({period})") as step:
                try:
                    recon_stats = _reconcile_single(Path(export_date_dir), statement_path, dry_run=False, console=console, quiet=True)
                    recon_stats_all.append(recon_stats)
                    all_recon_matches.extend(recon_stats.get("matches", []))
                    reconciled = recon_stats["reconciled"]
                    total = recon_stats["total"]
                    pct = recon_stats["reconciliation_rate"]
                    if total > 0:
                        step.success(f"{reconciled}/{total} reconciled ({pct:.0f}%)")
                    else:
                        step.warning("No transactions found")
                except Exception as exc:
                    step.warning(f"Reconciliation failed: {exc}")
                    logger.warning(f"Reconciliation failed for {statement_path.name}: {exc}")

        if merge_rules and all_recon_matches:
            with console.step_progress(f"Merge attachments ({export_date})") as step:
                try:
                    merge_stats = merge_reconciled_attachments(Path(export_date_dir), all_recon_matches, merge_rules)
                    merged = merge_stats["merged"]
                    errors = merge_stats["errors"]
                    if merged > 0:
                        step.success(f"{merged} document(s) merged")
                    elif errors > 0:
                        step.warning(f"{errors} merge error(s)")
                    else:
                        step.success("No merges needed")
                except Exception as exc:
                    step.warning(f"Merge attachments failed: {exc}")
                    logger.warning(f"Merge attachments failed: {exc}")

        with console.step_progress(f"Merge PDFs ({export_date})") as step:
            try:
                from pdf_gluer import merge_all_pdfs

                with suppress_console_logging():
                    merge_all_pdfs(export_date_dir)

                merged_files = list(Path(export_date_dir).glob("merged_*.pdf"))
                if merged_files:
                    prefixes = sorted({f.stem.split("_", 1)[1].upper() + "_" for f in merged_files if "_" in f.stem})
                    step.success(f"{len(merged_files)} merged PDFs ({', '.join(prefixes)})")
                else:
                    step.success("0 merged PDFs")
            except Exception as exc:
                step.error(f"PDF merge failed: {exc}")
                logger.error(f"Merge PDFs failed: {exc}")
                sys.exit(1)

        with suppress_console_logging():
            validate_merged_pdf(Path(export_date_dir))

        output_paths.append(("Export", str(export_date_dir)))

    if recon_stats_all:
        total_statements = len(recon_stats_all)
        total_reconciled = sum(s["reconciled"] for s in recon_stats_all)
        total_txns = sum(s["total"] for s in recon_stats_all)
        if total_txns > 0:
            avg_rate = total_reconciled / total_txns * 100
            summary["Reconciled"] = f"{total_statements} statement{'s' if total_statements != 1 else ''}, {avg_rate:.0f}% reconciled ({total_reconciled}/{total_txns})"

    output_paths.append(("Excel", str(processed_files_excel_path)))
    output_paths.append(("Log", str(log_file_path)))
    elapsed = time.time() - start_time
    console.pipeline_footer(
        elapsed_seconds=elapsed,
        warnings=pipeline_warnings if pipeline_warnings else None,
        summary=summary if summary else None,
        output_paths=output_paths,
    )
    logger.debug("All steps completed successfully.")


task_extract_new = extract
task_sync = sync
task_rename_files = rename
copy_matching_files = copy_matching
export_metadata_to_excel = export_excel
task_export_all_dates = export_dates
task_archive = archive
task_gmail_download = gmail
task_check = check
task_reconcile = reconcile


__all__ = [
    "archive",
    "check",
    "copy_matching",
    "copy_matching_files",
    "export_dates",
    "export_excel",
    "export_metadata_to_excel",
    "extract",
    "gmail",
    "pipeline",
    "reconcile",
    "rename",
    "sync",
    "task_archive",
    "task_check",
    "task_export_all_dates",
    "task_extract_new",
    "task_gmail_download",
    "task_reconcile",
    "task_rename_files",
    "task_sync",
    "validate_merged_pdf",
]
