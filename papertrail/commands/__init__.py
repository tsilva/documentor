"""Canonical command layer for papertrail."""

from __future__ import annotations

import fcntl
import hashlib
import io
import json
import re
import shutil
import sys
import tempfile
import time
import warnings
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from pathlib import Path
from typing import Optional

import fitz
import pandas as pd

from papertrail.archive_extract import extract_archives
from papertrail.config import (
    ExportSettings,
    NamingSettings,
    get_gmail_config_paths,
    get_passwords_from_profile,
)
from papertrail.document_types import normalize_document_type
from papertrail.engine import DocumentEngine
from papertrail.filename_audit import collect_long_filenames, format_long_filename_warning
from papertrail.gmail import download_gmail_attachments
from papertrail.logging_utils import (
    get_logger,
    setup_failure_logger,
    setup_task_logging,
    suppress_console_logging,
)
from papertrail.mbox import extract_mbox_attachments
from papertrail.models import DocumentMetadata, clean_enum_string
from papertrail.naming import sanitize_filename_component, trim_filename_component
from papertrail.pdf_merge import merge_all_pdfs
from papertrail.reconciliation_groundtruth import (
    GROUNDTRUTH_SUFFIX,
    document_hash_identity_matches,
)
from papertrail.repository import DocumentRepository
from papertrail.rules import RuleEngine
from papertrail.runtime import Runtime
from papertrail.utils import compute_month_range, make_matcher, month_to_date_range

from .check import run_check, validate_merged_pdf
from .reconcile import (
    discover_bank_statements,
    reconcile_single,
)

logger = get_logger("commands")
_PDFPRESS_COMPRESSORS = {}
_EXPORT_FILENAME_OMITTED_FIELDS = {"document_title"}


@contextmanager
def task_log_context(
    runtime: Runtime,
    processed_path: Path,
    task_name: str,
    *,
    show_header: bool = True,
):
    log_file_path = setup_task_logging(processed_path, task_name)
    logger.debug(f"=== {task_name.upper()} STARTED ===")
    logger.debug(f"Log: {log_file_path}")
    if show_header:
        runtime.console.detail(f"Log: {log_file_path}", indent=False)
    yield log_file_path


def _pipeline_step(console, label: str, action):
    with console.step_progress(label) as step:
        try:
            return action(step)
        except Exception as exc:
            step.error(str(exc))
            raise


def _load_json_dict(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as handle:
        data = json.load(handle)
    return data if isinstance(data, dict) else {}


def _statement_groundtruth_base_name(groundtruth_path: Path) -> str:
    name = groundtruth_path.name
    return name[: -len(GROUNDTRUTH_SUFFIX)] if name.endswith(GROUNDTRUTH_SUFFIX) else groundtruth_path.stem


def _statement_identity_key(document_name: str | None) -> tuple[str, str, str] | None:
    """Best-effort identity for statement approvals when duplicate XLSX bytes change."""
    if not document_name:
        return None
    stem = Path(document_name).stem
    match = re.search(r"\d{4}-\d{2}-\d{2}", stem)
    if not match:
        return None
    parts = stem[match.start() :].split(" - ")
    if len(parts) < 3:
        return None
    issuer_key = re.sub(r"[^a-z0-9]+", "", parts[2].lower())
    return parts[0], parts[1].lower(), issuer_key


def _snapshot_reconciliation_groundtruth(export_date_dir: Path) -> list[dict]:
    snapshots = []
    if not export_date_dir.exists():
        return snapshots

    for groundtruth_path in export_date_dir.rglob(f"*{GROUNDTRUTH_SUFFIX}"):
        base_name = _statement_groundtruth_base_name(groundtruth_path)
        metadata_path = groundtruth_path.with_name(f"{base_name}.json")
        metadata = {}
        if metadata_path.exists():
            try:
                metadata = _load_json_dict(metadata_path)
            except (OSError, UnicodeDecodeError, json.JSONDecodeError):
                metadata = {}

        try:
            payload = _load_json_dict(groundtruth_path)
        except (OSError, UnicodeDecodeError, json.JSONDecodeError):
            logger.warning(f"Could not preserve reconciliation approvals from {groundtruth_path.name}")
            continue

        snapshots.append(
            {
                "document_name": f"{base_name}.xlsx",
                "hash_file": metadata.get("hash_file"),
                "hash_content": metadata.get("hash_content"),
                "payload": payload,
            }
        )
    return snapshots


def _reconciliation_groundtruth_backup_path(export_dir: Path, export_date: str) -> Path:
    return export_dir / "_reconciliation_groundtruth" / f"{export_date}.json"


def _load_reconciliation_groundtruth_backup(export_dir: Path, export_date: str) -> list[dict]:
    backup_path = _reconciliation_groundtruth_backup_path(export_dir, export_date)
    if not backup_path.exists():
        return []
    try:
        data = json.loads(backup_path.read_text(encoding="utf-8"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError):
        logger.warning(f"Could not read reconciliation approval backup {backup_path}")
        return []
    return data if isinstance(data, list) else []


def _save_reconciliation_groundtruth_backup(
    export_dir: Path,
    export_date: str,
    snapshots: list[dict],
) -> None:
    if not snapshots:
        return
    backup_path = _reconciliation_groundtruth_backup_path(export_dir, export_date)
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path.write_text(
        json.dumps(snapshots, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _restore_reconciliation_groundtruth(
    snapshots: list[dict],
    bank_statements: list[Path],
) -> int:
    if not snapshots:
        return 0

    by_hash = {}
    by_name = {}
    by_statement_identity = {}
    for snapshot in snapshots:
        for key in (snapshot.get("hash_file"), snapshot.get("hash_content")):
            if key:
                by_hash[str(key)] = snapshot
        by_name[snapshot.get("document_name")] = snapshot
        identity_key = _statement_identity_key(snapshot.get("document_name"))
        if identity_key:
            by_statement_identity[identity_key] = snapshot

    restored = 0
    for statement_path in bank_statements:
        target_path = statement_path.with_suffix(GROUNDTRUTH_SUFFIX)
        if target_path.exists():
            continue

        metadata = {}
        metadata_path = statement_path.with_suffix(".json")
        if metadata_path.exists():
            try:
                metadata = _load_json_dict(metadata_path)
            except (OSError, UnicodeDecodeError, json.JSONDecodeError):
                metadata = {}

        snapshot = None
        for key in (metadata.get("hash_file"), metadata.get("hash_content")):
            if key and str(key) in by_hash:
                snapshot = by_hash[str(key)]
                break
        if snapshot is None:
            snapshot = by_name.get(statement_path.name)
        if snapshot is None:
            snapshot = by_statement_identity.get(_statement_identity_key(statement_path.name))
        if snapshot is None:
            continue

        payload = dict(snapshot["payload"])
        payload["source"] = statement_path.name
        for approval in payload.get("approvals", []):
            approval.setdefault("source_hint", {})["statement_file"] = statement_path.name
        for approval in payload.get("unmatched_file_approvals", []):
            approval.setdefault("source_hint", {})["statement_file"] = statement_path.name

        target_path.write_text(
            json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        restored += 1

    return restored


def _groundtruth_documents(snapshots: list[dict]) -> list[dict]:
    documents = []
    seen: set[tuple[str | None, str | None]] = set()
    for snapshot in snapshots:
        payload = snapshot.get("payload") or {}
        for approval in payload.get("approvals", []):
            for document in approval.get("required_documents", []):
                key = (document.get("hash_file"), document.get("hash_content"))
                if key in seen:
                    continue
                seen.add(key)
                documents.append(document)
        for approval in payload.get("unmatched_file_approvals", []):
            document = approval.get("document") or {}
            key = (document.get("hash_file"), document.get("hash_content"))
            if key in seen:
                continue
            seen.add(key)
            documents.append(document)
    return documents


def _export_contains_document(
    repository: DocumentRepository,
    export_date_dir: Path,
    identity: dict,
) -> bool:
    return any(
        document_hash_identity_matches(data, identity)
        for _, data in repository.iter_sidecars(export_date_dir)
    )


def _find_processed_document_by_identity(
    repository: DocumentRepository,
    processed_path: Path,
    identity: dict,
) -> tuple[Path, dict] | None:
    for json_path, data in repository.iter_sidecars(processed_path):
        if not document_hash_identity_matches(data, identity):
            continue
        doc_path = repository.find_companion(json_path, data)
        if doc_path is not None:
            return doc_path, data
    return None


def _profile_context(runtime: Runtime) -> dict | None:
    tax_number = runtime.profile.profile.tax_number
    return {"tax_number": tax_number} if tax_number else None


def _export_destination_paths(
    doc_path: Path,
    export_metadata_dict: dict,
    dest_folder: Path,
    *,
    export_config: ExportSettings,
    profile_context: Optional[dict],
    naming_settings: NamingSettings,
) -> tuple[Path, Path]:
    file_mappings = export_config.file_mappings
    engine = RuleEngine(profile_context=profile_context)

    if file_mappings.enabled:
        prefix = _sanitize_export_prefix(
            engine.evaluate_export_prefix(export_metadata_dict, file_mappings=file_mappings)
        )
        if file_mappings.filename_fields:
            file_hash = export_metadata_dict.get("hash_file", doc_path.stem.split(" - ")[-1])
            base_name = _build_filename_from_fields(
                export_metadata_dict,
                list(file_mappings.filename_fields),
                file_hash,
                profile_context=profile_context,
                component_max_chars=naming_settings.component_max_chars,
            )
        else:
            base_name = _sanitize_export_filename(doc_path.name)
        dest_name = _fit_pdf_export_filename(
            f"{prefix}{base_name}",
            max_length=naming_settings.pdf_export_max_chars,
        )
    else:
        dest_name = _fit_pdf_export_filename(
            _sanitize_export_filename(doc_path.name),
            max_length=naming_settings.pdf_export_max_chars,
        )

    dest_doc = dest_folder / dest_name
    return dest_doc, dest_doc.with_suffix(".json")


def _copy_document_to_export(
    repository: DocumentRepository,
    doc_path: Path,
    metadata_dict: dict,
    dest_folder: Path,
    *,
    export_config: ExportSettings,
    profile_context: Optional[dict],
    document_type_overrides=None,
    naming_settings: NamingSettings,
    max_file_size_mb: float | None,
) -> Path:
    export_metadata_dict = _metadata_for_export(
        metadata_dict,
        doc_path,
        document_type_overrides=document_type_overrides,
    )
    dest_doc, dest_json = _export_destination_paths(
        doc_path,
        export_metadata_dict,
        dest_folder,
        export_config=export_config,
        profile_context=profile_context,
        naming_settings=naming_settings,
    )
    _check_file_size(doc_path, max_file_size_mb)
    shutil.copy2(doc_path, dest_doc)
    if dest_doc.suffix.lower() == ".pdf":
        _compress_pdf_export(dest_doc, export_config=export_config)
    metadata_copy = dict(export_metadata_dict)
    metadata_copy["source_filename"] = doc_path.name
    repository.save_json(dest_json, metadata_copy)
    return dest_doc


def _restore_groundtruth_documents(
    runtime: Runtime,
    repository: DocumentRepository,
    processed_path: Path,
    export_date_dir: Path,
    snapshots: list[dict],
) -> int:
    export_config = runtime.profile.export
    restored = 0
    for identity in _groundtruth_documents(snapshots):
        if not identity or _export_contains_document(repository, export_date_dir, identity):
            continue
        found = _find_processed_document_by_identity(repository, processed_path, identity)
        if found is None:
            logger.warning(f"Could not restore approved export document {identity}")
            continue
        doc_path, metadata = found
        _copy_document_to_export(
            repository,
            doc_path,
            metadata,
            export_date_dir,
            export_config=export_config,
            profile_context=_profile_context(runtime),
            document_type_overrides=runtime.profile.classification.document_type_overrides,
            naming_settings=runtime.profile.naming,
            max_file_size_mb=export_config.max_file_size_mb,
        )
        restored += 1
    return restored


def _regenerate_export_period(
    runtime: Runtime,
    repository: DocumentRepository,
    processed_path: Path,
    export_dir: Path,
    export_date: str,
    *,
    step=None,
) -> tuple[Path, dict, list[Path]]:
    """Rebuild one export month while preserving its durable approvals."""
    export_date_dir = export_dir / export_date
    groundtruth_snapshots = _snapshot_reconciliation_groundtruth(export_date_dir)
    if groundtruth_snapshots:
        _save_reconciliation_groundtruth_backup(export_dir, export_date, groundtruth_snapshots)
    else:
        groundtruth_snapshots = _load_reconciliation_groundtruth_backup(export_dir, export_date)
    if export_date_dir.exists():
        shutil.rmtree(export_date_dir)

    copy_stats = copy_matching(
        runtime,
        processed_path,
        export_date,
        export_date_dir,
        incremental=False,
        quiet=True,
    )
    if step is not None:
        message = f"{copy_stats['copied']} files"
        if copy_stats.get("deduped", 0) > 0:
            message += f" ({copy_stats['deduped']} content dupes skipped)"
        step.success(message)

    restored_documents = _restore_groundtruth_documents(
        runtime,
        repository,
        processed_path,
        export_date_dir,
        groundtruth_snapshots,
    )
    if restored_documents:
        logger.debug(
            f"[RECON-GROUNDTRUTH] Restored {restored_documents} approved document(s) "
            f"for {export_date}"
        )

    bank_statements = discover_bank_statements(repository, export_date_dir)
    restored_groundtruth = _restore_reconciliation_groundtruth(
        groundtruth_snapshots,
        bank_statements,
    )
    if restored_groundtruth:
        logger.debug(
            f"[RECON-GROUNDTRUTH] Restored {restored_groundtruth} approval sidecar(s) "
            f"for {export_date}"
        )
    return export_date_dir, copy_stats, bank_statements


def _run_export_period(
    runtime: Runtime,
    processed_path: Path,
    export_dir: Path,
    export_date: str,
    *,
    merge_rules,
    run_merge_pdfs: bool = False,
) -> tuple[Path, list[dict]]:
    console = runtime.console

    def export_documents(step):
        return _regenerate_export_period(
            runtime,
            repository,
            processed_path,
            export_dir,
            export_date,
            step=step,
        )

    repository = DocumentRepository(runtime)
    export_date_dir, _, bank_statements = _pipeline_step(
        console,
        f"Export documents ({export_date})",
        export_documents,
    )
    all_recon_matches = []
    recon_stats_all: list[dict] = []

    for statement_path in bank_statements:
        statement_info = {}
        statement_json = statement_path.with_suffix(".json")
        if statement_json.exists():
            try:
                statement_info = repository.load_metadata(statement_json).get("bank_statement", {}) or {}
            except (OSError, UnicodeDecodeError, ValueError):
                pass
        account = statement_info.get("account_number", statement_path.stem)
        period = statement_info.get("period_start", export_date)
        if period and len(period) >= 7:
            period = period[:7]
        with console.step_progress(f"Match bank transactions: {account} ({period})") as step:
            try:
                recon_stats = reconcile_single(
                    runtime,
                    repository,
                    export_date_dir,
                    statement_path,
                    dry_run=False,
                    quiet=True,
                )
                recon_stats_all.append(recon_stats)
                all_recon_matches.extend(recon_stats.get("matches", []))
                total = recon_stats["total"]
                if total > 0:
                    step.success(
                        f"{recon_stats['reconciled']}/{total} reconciled "
                        f"({recon_stats['reconciliation_rate']:.0f}%)"
                    )
                else:
                    step.warning("No transactions found")
            except Exception as exc:
                step.warning(f"Reconciliation failed: {exc}")
                logger.warning(f"Reconciliation failed for {statement_path.name}: {exc}")

    if merge_rules and all_recon_matches:
        with console.step_progress(f"Merge attachments ({export_date})") as step:
            try:
                merge_stats = merge_reconciled_attachments(
                    runtime,
                    export_date_dir,
                    all_recon_matches,
                    merge_rules,
                )
                if merge_stats["merged"] > 0:
                    step.success(f"{merge_stats['merged']} document(s) merged")
                elif merge_stats["errors"] > 0:
                    step.warning(f"{merge_stats['errors']} merge error(s)")
                else:
                    step.success("No merges needed")
            except Exception as exc:
                step.warning(f"Merge attachments failed: {exc}")
                logger.warning(f"Merge attachments failed: {exc}")

    if run_merge_pdfs:
        def merge_pdfs(step):
            with suppress_console_logging():
                merged_outputs = merge_all_pdfs(str(export_date_dir))
                _compress_exported_pdfs(
                    list(merged_outputs.values()),
                    export_config=runtime.profile.export,
                )
            merged_files = list(export_date_dir.glob("merged_*.pdf"))
            if merged_files:
                prefixes = sorted(
                    {
                        file.stem.split("_", 1)[1].upper() + "_"
                        for file in merged_files
                        if "_" in file.stem
                    }
                )
                step.success(f"{len(merged_files)} merged PDFs ({', '.join(prefixes)})")
            else:
                step.success("0 merged PDFs")

        _pipeline_step(console, f"Merge PDFs ({export_date})", merge_pdfs)

        with suppress_console_logging():
            validate_merged_pdf(export_date_dir)

    return export_date_dir, recon_stats_all


def extract(
    runtime: Runtime,
    processed_path: Path,
    raw_paths: list[Path],
    *,
    quiet: bool = False,
) -> dict | None:
    lock_path = runtime.paths.cache / ".extract.lock"
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    lock_file = open(lock_path, "w")
    try:
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError:
        logger.error("Another extract process is already running. Exiting.")
        lock_file.close()
        return None

    try:
        with task_log_context(runtime, processed_path, "extract_new", show_header=not quiet):
            logs_dir = processed_path / "logs"
            failure_log_path = logs_dir / "classification_failures.log"
            failure_logger = setup_failure_logger(failure_log_path)
            return DocumentEngine(runtime).extract(
                processed_path,
                raw_paths,
                quiet=quiet,
                failure_logger=failure_logger,
            )
    finally:
        fcntl.flock(lock_file, fcntl.LOCK_UN)
        lock_file.close()


def sync(
    runtime: Runtime,
    processed_path: Path,
    *,
    dry_run: bool = False,
    all_unknown: bool = False,
    pattern: str | None = None,
    workers: int = 1,
    all: bool = False,
    quiet: bool = False,
) -> dict:
    with task_log_context(runtime, processed_path, "sync", show_header=not quiet):
        return DocumentEngine(runtime).sync(
            processed_path,
            dry_run=dry_run,
            all_unknown=all_unknown,
            pattern=pattern,
            workers=workers,
            all=all,
            quiet=quiet,
        )


def rename(runtime: Runtime, processed_path: Path, *, quiet: bool = False) -> dict:
    export_root = runtime.paths.export
    if export_root:
        export_path = export_root.resolve()
        target_path = processed_path.resolve()
        if target_path == export_path or export_path in target_path.parents:
            raise RuntimeError(
                "rename cannot be run on export directories; regenerate exports instead."
            )

    repository = DocumentRepository(runtime)
    with task_log_context(runtime, processed_path, "rename_files", show_header=not quiet):
        stats = repository.repair_filenames(processed_path)
        if not quiet:
            runtime.console.success(
                f"{stats['validated']} files validated, {stats['renamed']} renamed",
                indent=False,
            )
        logger.debug(f"Renaming complete. Renamed {stats['renamed']} files.")
        return stats


def _truncate_filename_component(value: str, max_length: int) -> str:
    if max_length <= 0:
        return ""
    if len(value) <= max_length:
        return value
    return value[:max_length].rstrip(" -_.") or value[:max_length]


def _fit_pdf_export_filename(filename: str, max_length: int) -> str:
    if not filename.lower().endswith(".pdf") or len(filename) <= max_length:
        return filename

    suffix = ".pdf"
    parts = filename[: -len(suffix)].split(" - ")
    if len(parts) < 4:
        return _truncate_filename_component(filename[: -len(suffix)], max_length - len(suffix)) + suffix
    if len(parts) > 4:
        parts = parts[:3] + parts[-1:]

    def candidate_name() -> str:
        return " - ".join(parts) + suffix

    for index in (2, 1):
        overage = len(candidate_name()) - max_length
        if overage <= 0:
            break
        available = max(len(parts[index]) - 1, 0)
        if available <= 0:
            continue
        parts[index] = _truncate_filename_component(parts[index], len(parts[index]) - min(overage, available))

    candidate = candidate_name()
    if len(candidate) <= max_length:
        return candidate

    overage = len(candidate) - max_length
    parts[0] = _truncate_filename_component(parts[0], max(len(parts[0]) - overage, 1))
    return candidate_name()


def _build_filename_from_fields(
    metadata: dict,
    fields: list[str],
    file_hash: str,
    *,
    profile_context: dict | None = None,
    component_max_chars: int,
) -> str:
    engine = RuleEngine(profile_context=profile_context)
    parts = []
    for field_name in fields:
        if field_name in _EXPORT_FILENAME_OMITTED_FIELDS:
            continue
        value = engine.get_nested_value(metadata, field_name)
        if value is not None and str(value).strip():
            component = engine.resolve_profile_value(str(value))
            component = sanitize_filename_component(component.strip())
            component = trim_filename_component(component, component_max_chars)
            parts.append(component)
    extension = metadata.get("source_extension") or ".pdf"
    parts.append(f"{file_hash}{extension}")
    return " - ".join(parts).lower()


def _sanitize_export_prefix(prefix: str) -> str:
    return sanitize_filename_component(prefix)


def _sanitize_export_filename(filename: str) -> str:
    suffix = Path(filename).suffix
    stem = filename[: -len(suffix)] if suffix else filename
    clean_stem = sanitize_filename_component(stem)
    clean_suffix = sanitize_filename_component(suffix).lower()
    return f"{clean_stem}{clean_suffix}"


def _should_skip_copy(src: Path, dst: Path) -> bool:
    from papertrail.hashing import hash_file_fast

    return dst.exists() and src.stat().st_size == dst.stat().st_size and hash_file_fast(src) == hash_file_fast(dst)


def _check_file_size(src: Path, max_file_size_mb: float | None) -> None:
    if max_file_size_mb is None:
        return
    size = src.stat().st_size
    threshold = max_file_size_mb * 1024 * 1024
    if size >= threshold:
        size_mb = size / (1024 * 1024)
        logger.warning(f"Large file: {src.name} ({size_mb:.1f} MB exceeds {max_file_size_mb} MB threshold)")


def _compression_enabled(pdf_path: Path, export_config: ExportSettings) -> bool:
    settings = export_config.compression
    if not settings.enabled:
        return False
    min_size_mb = settings.min_size_mb
    if min_size_mb is not None and pdf_path.stat().st_size < float(min_size_mb) * 1024 * 1024:
        return False
    return True


def _get_pdfpress_compressor(export_config: ExportSettings):
    quality = export_config.compression.quality
    if quality not in _PDFPRESS_COMPRESSORS:
        try:
            from pdfpress import PDFCompressor
        except ImportError as exc:
            raise RuntimeError(
                "pdfpress is required for export PDF compression. Install the project dependencies again."
            ) from exc
        _PDFPRESS_COMPRESSORS[quality] = PDFCompressor(quality=quality)
    return _PDFPRESS_COMPRESSORS[quality]


def _compress_pdf_export(pdf_path: Path, *, export_config: ExportSettings) -> None:
    if pdf_path.suffix.lower() != ".pdf" or not pdf_path.exists():
        return
    if not _compression_enabled(pdf_path, export_config):
        return

    original_size = pdf_path.stat().st_size
    compressor = _get_pdfpress_compressor(export_config)

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_output = Path(tmp_dir) / pdf_path.name
        outcome = compressor.compress(pdf_path, tmp_output)
        shutil.move(tmp_output, pdf_path)

    final_size = pdf_path.stat().st_size
    logger.debug(
        f"[EXPORT-COMPRESS] {pdf_path.name}: {original_size} -> {final_size} bytes "
        f"(strategy={outcome.best_strategy})"
    )


def _compress_exported_pdfs(pdf_paths: list[Path], *, export_config: ExportSettings) -> None:
    for pdf_path in pdf_paths:
        _compress_pdf_export(pdf_path, export_config=export_config)


def _long_filename_warning(path: Path, *, naming_settings: NamingSettings) -> str:
    return format_long_filename_warning(
        collect_long_filenames(
            path,
            max_length=naming_settings.filename_warning_max_chars,
            suffixes=(".pdf",),
        ),
        max_items=None,
    )


def _warn_long_filenames(runtime: Runtime, path: Path) -> str:
    warning = _long_filename_warning(path, naming_settings=runtime.profile.naming)
    if not warning:
        return ""
    runtime.console.warning(warning, indent=False)
    return warning


def _metadata_for_export(metadata: dict, doc_path: Path, *, document_type_overrides=None) -> dict:
    export_metadata = dict(metadata)
    if doc_path.suffix.lower() == ".pdf":
        effective_type = normalize_document_type(
            export_metadata.get("document_type"),
            export_metadata.get("document_type_raw"),
            export_metadata.get("document_title"),
            document_type_overrides,
        )
        if effective_type:
            export_metadata["document_type"] = effective_type
    return export_metadata


def _export_dedup_key(metadata: dict) -> tuple[str, ...] | None:
    bank_statement = metadata.get("bank_statement")
    if isinstance(bank_statement, dict):
        account_number = bank_statement.get("account_number")
        period_start = bank_statement.get("period_start")
        period_end = bank_statement.get("period_end")
        if account_number and period_start and period_end:
            return (
                "bank-statement",
                str(bank_statement.get("bank_format") or ""),
                str(account_number),
                str(period_start),
                str(period_end),
            )

    content_hash = metadata.get("hash_content")
    if content_hash:
        return ("content", str(content_hash))
    return None


def copy_matching(
    runtime: Runtime,
    processed_path: Path,
    pattern: str,
    dest_folder: Path,
    *,
    incremental: bool = False,
    quiet: bool = False,
) -> dict:
    repository = DocumentRepository(runtime)
    export_config = runtime.profile.export
    profile_context = _profile_context(runtime)
    matcher = make_matcher(pattern, use_search=True)
    dest_folder.mkdir(parents=True, exist_ok=True)

    max_file_size_mb = export_config.max_file_size_mb

    stats = {"copied": 0, "skipped": 0, "deduped": 0, "total": 0}
    seen_dedup_keys: set[tuple[str, ...]] = set()

    documents = sorted(
        repository.iter_documents(processed_path, validate=False, require_companion=True),
        key=lambda item: str(item[0]),
    )
    for json_path, doc_path, metadata in runtime.console.track(documents, "Copying files"):
        metadata_dict = metadata.model_dump() if isinstance(metadata, DocumentMetadata) else metadata
        export_metadata_dict = _metadata_for_export(
            metadata_dict,
            doc_path,
            document_type_overrides=runtime.profile.classification.document_type_overrides,
        )
        if not matcher(doc_path.name) and not matcher(json_path.name):
            continue

        stats["total"] += 1
        dedup_key = _export_dedup_key(export_metadata_dict)
        if dedup_key and dedup_key in seen_dedup_keys:
            stats["deduped"] += 1
            logger.debug(f"[EXPORT-DEDUP] Skipping {doc_path.name} (dedup key {dedup_key} already exported)")
            continue
        if dedup_key:
            seen_dedup_keys.add(dedup_key)

        dest_doc, dest_json = _export_destination_paths(
            doc_path,
            export_metadata_dict,
            dest_folder,
            export_config=export_config,
            profile_context=profile_context,
            naming_settings=runtime.profile.naming,
        )

        if incremental and _should_skip_copy(doc_path, dest_doc):
            stats["skipped"] += 1
            continue

        _check_file_size(doc_path, max_file_size_mb)
        shutil.copy2(doc_path, dest_doc)
        if dest_doc.suffix.lower() == ".pdf":
            _compress_pdf_export(dest_doc, export_config=export_config)
        metadata_copy = dict(export_metadata_dict)
        metadata_copy["source_filename"] = doc_path.name
        repository.save_json(dest_json, metadata_copy)
        stats["copied"] += 1

    if not quiet:
        runtime.console.success(f"Copied {stats['copied']} files to {dest_folder.name}", indent=False)
        _warn_long_filenames(runtime, dest_folder)
    return stats


def export_excel(runtime: Runtime, processed_path: Path, excel_output_path: str, *, quiet: bool = False) -> dict:
    repository = DocumentRepository(runtime)
    metadata_list = []

    for metadata_path, metadata in repository.load_sidecars_parallel(
        processed_path,
        validate=True,
        show_progress=not quiet,
        progress_desc="Collecting metadata",
    ):
        metadata_dict = metadata.model_dump()
        metadata_dict.pop("class_reasoning", None)

        doc_path = repository.find_companion(metadata_path, metadata_dict)
        filename = doc_path.name if doc_path and doc_path.exists() else ""
        metadata_dict["filename"] = filename
        metadata_dict["filename_length"] = len(filename)

        try:
            date_parts = metadata.date_issued.split("-")
            metadata_dict["year"] = int(date_parts[0])
            metadata_dict["month"] = int(date_parts[1])
        except (IndexError, ValueError, AttributeError):
            metadata_dict["year"] = None
            metadata_dict["month"] = None

        if isinstance(metadata_dict.get("document_type"), str):
            metadata_dict["document_type"] = clean_enum_string(metadata_dict["document_type"], "DocumentType")

        metadata_list.append(metadata_dict)

    if not metadata_list:
        if not quiet:
            runtime.console.warning("No valid metadata found to export", indent=False)
        return {"exported": 0}

    dataframe = pd.DataFrame(metadata_list)
    ordered_cols = [
        "class_confidence",
        "date_issued",
        "year",
        "month",
        "hash_content",
        "hash_file",
        "filename",
        "filename_length",
        "page_count",
        "document_type",
        "document_type_raw",
        "document_title",
        "issuing_party",
        "issuing_party_raw",
        "total_amount",
        "total_amount_currency",
    ]
    extra_cols = [col for col in dataframe.columns if col not in ordered_cols]
    dataframe = dataframe[ordered_cols + extra_cols]

    if "date_issued" in dataframe.columns:
        dataframe = dataframe.sort_values(by="date_issued", ascending=False)

    with pd.ExcelWriter(excel_output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.sheets["Sheet1"]
        worksheet.freeze_panes = "A2"

        from openpyxl.utils import get_column_letter

        for col in ordered_cols:
            if col in dataframe.columns:
                col_idx = dataframe.columns.get_loc(col) + 1
                col_letter = get_column_letter(col_idx)
                value_lengths = [len(str(value)) for value in dataframe[col].values if value is not None]
                max_len = max(value_lengths + [len(col)])
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 102)

        for col in ("year", "month", "filename_length"):
            if col in dataframe.columns:
                col_letter = get_column_letter(dataframe.columns.get_loc(col) + 1)
                worksheet.column_dimensions[col_letter].hidden = True

    if not quiet:
        runtime.console.success(f"Exported {len(dataframe)} entries", indent=False)
    logger.debug(f"Exported {len(dataframe)} entries to {excel_output_path}")
    return {"exported": len(dataframe)}


def _calculate_directory_hash(directory: Path) -> str:
    from papertrail.hashing import hash_file_fast

    pdf_files = sorted(directory.glob("*.pdf"))
    if not pdf_files:
        return ""
    combined = [f"{path.name}:{hash_file_fast(path)}" for path in pdf_files]
    return hashlib.sha256("\n".join(combined).encode()).hexdigest()[:16]


def _directory_has_changed(directory: Path) -> bool:
    hash_file_path = directory / ".directory_hash"
    current_hash = _calculate_directory_hash(directory)
    if not current_hash:
        return False
    if not hash_file_path.exists():
        hash_file_path.write_text(current_hash, encoding="utf-8")
        return True
    stored_hash = hash_file_path.read_text(encoding="utf-8").strip()
    if current_hash != stored_hash:
        hash_file_path.write_text(current_hash, encoding="utf-8")
        return True
    return False


def export_dates(
    runtime: Runtime,
    processed_path: Path,
    export_base_dir: Path,
    run_merge: bool = False,
) -> None:
    repository = DocumentRepository(runtime)
    export_config = runtime.profile.export

    with task_log_context(runtime, processed_path, "export_all_dates", show_header=False):
        all_dates = repository.unique_dates(processed_path)
        if not all_dates:
            runtime.console.warning("No dates found in processed files", indent=False)
            return

        total_copied = 0
        total_skipped = 0
        changed_directories = []
        for date in runtime.console.track(all_dates, "Exporting dates"):
            export_date_dir, stats, _ = _regenerate_export_period(
                runtime,
                repository,
                processed_path,
                export_base_dir,
                date,
            )

            total_copied += stats["copied"]
            total_skipped += stats["skipped"]

            if stats["copied"] > 0:
                changed_directories.append(export_date_dir)
            elif stats["total"] > 0 and export_date_dir.exists() and _directory_has_changed(export_date_dir):
                changed_directories.append(export_date_dir)

        runtime.console.success(f"{len(all_dates)} dates exported, {total_copied} files copied", indent=False)
        logger.debug(
            f"Processed {len(all_dates)} date(s), Total files copied: {total_copied}, Skipped: {total_skipped}"
        )

        if run_merge and changed_directories:
            for export_dir in runtime.console.track(changed_directories, "Merging PDFs"):
                try:
                    merged_outputs = merge_all_pdfs(str(export_dir))
                    _compress_exported_pdfs(
                        list(merged_outputs.values()),
                        export_config=export_config,
                    )
                    validate_merged_pdf(export_dir)
                except Exception as exc:
                    logger.error(f"Merge failed: {exc}")

            runtime.console.success(f"Merged {len(changed_directories)} directories", indent=False)

        _warn_long_filenames(runtime, export_base_dir)


def archive(runtime: Runtime, processed_path: Path, digests: list[str], *, dry_run: bool = False) -> None:
    repository = DocumentRepository(runtime)
    if dry_run:
        runtime.console.info("Dry run - no files will be moved", indent=False)
    stats = repository.archive_by_hash_file(digests, dry_run=dry_run, scope=processed_path)
    for digest in stats["not_found"]:
        runtime.console.warning(f"[NOT FOUND] {digest}", indent=False)

    runtime.console.info(
        f"Archive: {stats['found']} found, {stats['archived']} archived, {len(stats['not_found'])} not found",
        indent=False,
    )
    if not dry_run and stats["archived"] > 0:
        runtime.console.detail(f"Archived to: {stats['archive_dir']}", indent=False)


def gmail(runtime: Runtime, *, months: int = 2) -> None:
    raw_paths = runtime.paths.raw
    processed_path = runtime.paths.processed
    gmail_settings = runtime.profile.gmail

    if processed_path:
        setup_task_logging(processed_path, "gmail_download")

    has_gmail_output = bool(raw_paths or gmail_settings.output_raw_path)
    if not has_gmail_output or not processed_path:
        missing = []
        if not has_gmail_output:
            missing.append("paths.raw or gmail.output_raw_path")
        if not processed_path:
            missing.append("paths.processed")
        runtime.console.error(f"Missing required profile settings: {', '.join(missing)}", indent=False)
        raise RuntimeError(f"Missing required profile settings: {', '.join(missing)}")

    raw_path = (
        Path(gmail_settings.output_raw_path)
        if gmail_settings.output_raw_path
        else raw_paths[0]
    )
    export_dates_list = compute_month_range(months)
    totals = {
        "messages_found": 0,
        "messages_processed": 0,
        "messages_skipped": 0,
        "attachments_downloaded": 0,
        "attachments_failed": 0,
        "bytes_downloaded": 0,
    }
    gmail_dir = raw_path / (gmail_settings.output_subdir or "gmail")
    tracking_dir = (
        Path(gmail_settings.tracking_dir)
        if gmail_settings.tracking_dir
        else processed_path
        / "logs"
        / (gmail_settings.tracking_subdir or "gmail_tracking")
    )
    tracking_dir.mkdir(parents=True, exist_ok=True)
    paths = get_gmail_config_paths(runtime.profile)

    for month in export_dates_list:
        month_dir = gmail_dir / month
        month_dir.mkdir(parents=True, exist_ok=True)
        start_date, end_date = month_to_date_range([month])
        logger.debug(f"Gmail {month}: {start_date.date()} to {end_date.date()} -> {month_dir}")

        try:
            stats = download_gmail_attachments(
                output_dir=month_dir,
                start_date=start_date,
                end_date=end_date,
                tracking_dir=tracking_dir,
                credentials_path=paths["credentials"],
                token_path=paths["token"],
                settings=runtime.profile.gmail,
                console=runtime.console,
            )
        except FileNotFoundError as exc:
            runtime.console.error(f"Gmail credentials not found: {exc}", indent=False)
            raise RuntimeError(f"Gmail credentials not found: {exc}") from exc
        except Exception as exc:
            error_type = type(exc).__name__
            runtime.console.error(f"Gmail download failed ({error_type}): {exc}", indent=False)
            raise RuntimeError(f"Gmail download failed ({error_type}): {exc}") from exc

        for key in totals:
            totals[key] += stats[key]

    if totals["attachments_downloaded"] > 0:
        runtime.console.success(
            f"{totals['messages_processed']} messages processed, {totals['attachments_downloaded']} new attachments",
            indent=False,
        )
    elif totals["messages_processed"] > 0:
        runtime.console.success(f"{totals['messages_processed']} messages processed, 0 new attachments", indent=False)
    else:
        date_range = f"{export_dates_list[0]} to {export_dates_list[-1]}"
        runtime.console.warning(f"No messages found ({date_range})", indent=False)


def check(runtime: Runtime, processed_path: Path, *, verify_hashes: bool = False, dry_run: bool = False) -> None:
    repository = DocumentRepository(runtime)
    engine = DocumentEngine(runtime, repository)
    run_check(runtime, repository, engine, processed_path, verify_hashes=verify_hashes, dry_run=dry_run)


def _page_signature(page: fitz.Page) -> str:
    text = page.get_text("text").strip()
    if text:
        payload = text.encode("utf-8")
    else:
        pix = page.get_pixmap(matrix=fitz.Matrix(0.5, 0.5), colorspace=fitz.csGRAY, alpha=False)
        payload = f"{pix.width}x{pix.height}:".encode("ascii") + bytes(pix.samples)
    return hashlib.sha256(payload).hexdigest()


def _pdf_tail_matches_attachment(target_pdf: Path, attach_pdf: Path) -> bool:
    with fitz.open(target_pdf) as target_doc, fitz.open(attach_pdf) as attach_doc:
        attach_page_count = len(attach_doc)
        if attach_page_count == 0 or len(target_doc) < attach_page_count:
            return False
        start_index = len(target_doc) - attach_page_count
        for offset in range(attach_page_count):
            if _page_signature(target_doc[start_index + offset]) != _page_signature(attach_doc[offset]):
                return False
    return True


def merge_reconciled_attachments(runtime: Runtime, export_path: Path, all_matches: list, merge_rules: list) -> dict:
    stats = {"merged": 0, "skipped": 0, "errors": 0}
    if not merge_rules or not all_matches:
        return stats

    merged_attachments: set[str] = set()
    modified_targets: set[Path] = set()
    engine = RuleEngine()

    for match in all_matches:
        for target, attachment in engine.select_merge_pairs(match, merge_rules):
            target_pdf = export_path / target.pdf_filename
            if not target_pdf.exists() or target_pdf.suffix.lower() != ".pdf":
                continue

            if attachment.pdf_filename in merged_attachments:
                stats["skipped"] += 1
                continue

            attach_pdf = export_path / attachment.pdf_filename
            if not attach_pdf.exists() or attach_pdf.suffix.lower() != ".pdf":
                continue
            if target_pdf == attach_pdf:
                continue
            if _pdf_tail_matches_attachment(target_pdf, attach_pdf):
                merged_attachments.add(attachment.pdf_filename)
                stats["skipped"] += 1
                continue

            try:
                import pikepdf

                with pikepdf.open(target_pdf, allow_overwriting_input=True) as target_doc:
                    with pikepdf.open(attach_pdf) as attach_doc:
                        target_doc.pages.extend(attach_doc.pages)
                    target_doc.save(target_pdf)
                merged_attachments.add(attachment.pdf_filename)
                modified_targets.add(target_pdf)
                stats["merged"] += 1
            except Exception as exc:
                stats["errors"] += 1
                logger.error(
                    f"[MERGE] Failed to append {attachment.pdf_filename} to {target.pdf_filename}: {exc}"
                )

    _compress_exported_pdfs(sorted(modified_targets), export_config=runtime.profile.export)

    return stats


def reconcile(
    runtime: Runtime,
    export_path: Path,
    *,
    excel_path: Optional[Path] = None,
    dry_run: bool = False,
) -> None:
    repository = DocumentRepository(runtime)
    merge_rules = runtime.profile.export.merge_rules
    with task_log_context(runtime, export_path, "reconcile"):
        if excel_path is not None:
            excel_paths = [excel_path]
        else:
            excel_paths = discover_bank_statements(repository, export_path)

        if not excel_paths:
            runtime.console.warning("No bank statements found to reconcile", indent=False)
            return

        all_recon_matches: list[dict] = []
        for path in excel_paths:
            if not path.exists():
                runtime.console.error(f"Excel file not found: {path}", indent=False)
                continue
            with runtime.console.task(f"Reconcile: {path.name}"):
                stats = reconcile_single(runtime, repository, export_path, path, dry_run=dry_run)
                all_recon_matches.extend(stats.get("matches", []))

        if not dry_run and merge_rules and all_recon_matches:
            merge_stats = merge_reconciled_attachments(runtime, export_path, all_recon_matches, merge_rules)
            if merge_stats["merged"] > 0:
                runtime.console.success(
                    f"Merged attachments into {merge_stats['merged']} document(s)",
                    indent=False,
                )
            elif merge_stats["errors"] > 0:
                runtime.console.warning(
                    f"Attachment merge completed with {merge_stats['errors']} error(s)",
                    indent=False,
                )


def review(runtime: Runtime, export_path: Path) -> None:
    from tools.review import launch

    launch(export_path=export_path, argv=[])


def pipeline(
    runtime: Runtime,
    *,
    months: int = 2,
    export_date_arg: Optional[str] = None,
) -> None:
    console = runtime.console
    start_time = time.time()
    profile = runtime.profile

    raw_dirs = runtime.paths.raw
    processed_dir = runtime.paths.processed
    export_dir = runtime.paths.export

    missing = []
    if not raw_dirs:
        missing.append("paths.raw")
    if not processed_dir:
        missing.append("paths.processed")
    if not export_dir:
        missing.append("paths.export")
    if missing:
        console.error(f"Missing required profile settings: {', '.join(missing)}", indent=False)
        sys.exit(1)

    processed_path = Path(processed_dir)
    log_file_path = setup_task_logging(processed_path, "pipeline")
    console.pipeline_header(profile.profile.name, str(log_file_path))

    export_dates_list = [export_date_arg] if export_date_arg else compute_month_range(months)
    for export_date in export_dates_list:
        if not re.match(r"^\d{4}-\d{2}$", export_date):
            console.error(f"The export_date must be in YYYY-MM format: {export_date}", indent=False)
            sys.exit(1)

    passwords, _ = get_passwords_from_profile(profile)
    if not passwords:
        logger.debug("No passwords configured. Password-protected archives will be skipped.")

    processed_files_excel_path = processed_path / "processed_files.xlsx"
    pipeline_warnings: list[str] = []
    summary: dict[str, str] = {}
    output_paths: list[tuple[str, str]] = []

    if profile.gmail.enabled:
        with console.step_progress("Download Gmail attachments") as step:
            try:
                gmail(runtime, months=months)
                step.success("Download completed")
            except Exception as exc:
                msg = f"Gmail download failed, continuing pipeline ({exc})"
                step.warning(msg)
                pipeline_warnings.append(msg)
                logger.warning(f"Gmail download failed (non-fatal): {exc}")

    for raw_dir in raw_dirs:
        with console.step_progress("Extract mbox attachments") as step:
            stats = extract_mbox_attachments(str(raw_dir))
            if stats["mbox_files"] > 0:
                step.success(f"{stats['mbox_files']} mbox file(s), {stats['attachments_extracted']} attachment(s)")
            else:
                step.warning("No mbox files found")
            if stats["errors"]:
                step.error(f"{len(stats['errors'])} error(s)")
                sys.exit(1)

        with console.step_progress("Extract compressed archives") as step:
            with redirect_stdout(io.StringIO()), redirect_stderr(io.StringIO()), warnings.catch_warnings():
                warnings.simplefilter("ignore")
                results = extract_archives(str(raw_dir), passwords=passwords if passwords else None)
            total_extracted = 0
            failures = 0
            for count in results.values():
                if count == -1:
                    failures += 1
                else:
                    total_extracted += count
            if total_extracted > 0:
                step.success(f"{total_extracted} files from {len(results) - failures} archive(s)")
            elif failures > 0:
                step.warning(f"{failures} archive(s) failed")
            else:
                step.warning("No archives found")

    console.step("Classify new documents")
    try:
        extract_stats = extract(runtime, processed_path, raw_dirs, quiet=False)
        if extract_stats is None:
            console.warning("Extraction locked by another process")
        elif extract_stats["failed"] > 0:
            console.error(f"LLM API error ({extract_stats['failed']} failed)")
    except Exception as exc:
        console.error(str(exc))
        sys.exit(1)

    if extract_stats and extract_stats.get("bundles_split", 0) > 0:
        console.notable(f"Split {extract_stats['bundles_split']} PDF bundles into {extract_stats['split_pages']} pages")
    if extract_stats:
        if extract_stats.get("new_issuers"):
            console.notable(f"New issuers: {', '.join(extract_stats['new_issuers'])}")
        unknown_parts = []
        if extract_stats.get("unknown_document_type"):
            unknown_parts.append(f"{extract_stats['unknown_document_type']} document_type")
        if extract_stats.get("unknown_issuing_party"):
            unknown_parts.append(f"{extract_stats['unknown_issuing_party']} issuing_party")
        if unknown_parts:
            unknown_msg = f"Unknown: {', '.join(unknown_parts)}"
            console.notable(unknown_msg)
            pipeline_warnings.append(f"{unknown_msg} in extracted documents")

        if extract_stats["new"] > 0 or extract_stats["duplicates"] > 0:
            classified_parts = [f"{extract_stats['new']} new", f"{extract_stats['duplicates']} duplicates"]
            if extract_stats.get("batch_duplicates", 0) > 0:
                classified_parts.append(f"{extract_stats['batch_duplicates']} batch dupes")
            summary["Classified"] = ", ".join(classified_parts)

    console.step("Sync orphaned metadata")
    try:
        sync_stats = sync(runtime, processed_path, quiet=False)
        orphans = sync_stats.get("targets", 0)
        if orphans == 0:
            console.success("0 orphans found")
        else:
            resynced = sync_stats.get("new", 0) + sync_stats.get("changed", 0)
            console.success(f"{resynced} orphans re-synced")
    except Exception as exc:
        console.error(str(exc))
        sys.exit(1)

    with console.step_progress("Sync filenames to metadata") as step:
        try:
            rename_stats = rename(runtime, processed_path, quiet=True)
            step.success(f"{rename_stats['validated']} validated, {rename_stats['renamed']} renamed")
            if rename_stats["renamed"] > 0:
                summary["Renamed"] = f"{rename_stats['renamed']} files"
            orphans = rename_stats.get("orphans", 0)
            if orphans > 0:
                msg = f"{orphans} orphaned JSON sidecar(s) (companion file missing)"
                console.notable(msg)
                pipeline_warnings.append(msg)
        except Exception as exc:
            step.error(str(exc))
            sys.exit(1)

    with console.step_progress("Export to Excel") as step:
        try:
            excel_stats = export_excel(runtime, processed_path, str(processed_files_excel_path), quiet=True)
            if excel_stats["exported"]:
                step.success(f"{excel_stats['exported']} entries")
                summary["Exported"] = f"{excel_stats['exported']} entries"
            else:
                step.warning("No valid metadata found to export")
        except Exception as exc:
            step.error(str(exc))
            sys.exit(1)

    recon_stats_all: list[dict] = []
    merge_rules = profile.export.merge_rules

    for export_date in export_dates_list:
        try:
            export_date_dir, period_recon_stats = _run_export_period(
                runtime,
                processed_path,
                Path(export_dir),
                export_date,
                merge_rules=merge_rules,
                run_merge_pdfs=False,
            )
        except Exception as exc:
            logger.error(f"Export pipeline failed for {export_date}: {exc}")
            sys.exit(1)

        recon_stats_all.extend(period_recon_stats)
        output_paths.append(("Export", str(export_date_dir)))

    if recon_stats_all:
        total_statements = len(recon_stats_all)
        total_reconciled = sum(stats["reconciled"] for stats in recon_stats_all)
        total_txns = sum(stats["total"] for stats in recon_stats_all)
        if total_txns > 0:
            avg_rate = total_reconciled / total_txns * 100
            summary["Reconciled"] = (
                f"{total_statements} statement{'s' if total_statements != 1 else ''}, "
                f"{avg_rate:.0f}% reconciled ({total_reconciled}/{total_txns})"
            )

    if filename_warning := _long_filename_warning(
        Path(export_dir),
        naming_settings=runtime.profile.naming,
    ):
        pipeline_warnings.append(filename_warning)

    output_paths.append(("Excel", str(processed_files_excel_path)))
    output_paths.append(("Log", str(log_file_path)))
    elapsed = time.time() - start_time
    console.pipeline_footer(
        elapsed_seconds=elapsed,
        warnings=pipeline_warnings if pipeline_warnings else None,
        summary=summary if summary else None,
        output_paths=output_paths,
    )
    logger.debug("All steps completed successfully.")


__all__ = [
    "archive",
    "check",
    "copy_matching",
    "export_dates",
    "export_excel",
    "extract",
    "gmail",
    "merge_reconciled_attachments",
    "pipeline",
    "reconcile",
    "review",
    "rename",
    "sync",
    "validate_merged_pdf",
]
