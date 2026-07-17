"""Millennium BCP bank statement parser."""

from pathlib import Path

import openpyxl

from papertrail.bank_statement.models import (
    BankFormat,
    BankStatementData,
    BankTransactionRecord,
    parse_bank_amount,
)
from papertrail.bank_statement.parser_utils import (
    cfg_cell,
    cfg_int,
    cfg_sequence,
    cfg_str,
    expected_headers,
    normalized_headers,
    parse_date_cell,
    parse_date_str,
)
from papertrail.logging_utils import get_logger

logger = get_logger("bank_statement")

FORMAT = BankFormat.MILLENNIUM_BCP

DEFAULT_CONFIG: dict[str, object] = {
    "header_row": 8,
    "data_start_row": 9,
    "scan_columns": 7,
    "expected_headers": ("data lancamento", "descricao", "montante"),
    "date_formats": ("%d/%m/%Y", "%d-%m-%Y"),
    "account_cell": (2, 3),
    "period_start_cell": (3, 3),
    "period_end_cell": (4, 3),
    "account_currency_separator": " - ",
    "default_currency": "EUR",
    "issuer_party": "MillenniumBCP",
    "issuer_party_raw": "Millennium BCP",
    "max_columns": 7,
    "description_column": 3,
    "amount_column": 4,
    "currency_column": 5,
    "notes_column": 6,
    "treated_column": 7,
    "untreated_values": ("nao", "não", ""),
}


def can_parse(ws, config: dict[str, object] | None = None) -> bool:
    """Detect Millennium BCP format by checking column headers in row 8."""
    header_row = cfg_int(config, "header_row", DEFAULT_CONFIG["header_row"])
    scan_columns = cfg_int(config, "scan_columns", DEFAULT_CONFIG["scan_columns"])
    return expected_headers(config, DEFAULT_CONFIG["expected_headers"]).issubset(
        normalized_headers(ws, row=header_row, columns=scan_columns)
    )


def parse(xlsx_path: Path, config: dict[str, object] | None = None) -> BankStatementData | None:
    """Parse a Millennium BCP bank statement XLSX."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    if not can_parse(ws, config=config):
        wb.close()
        return None

    account_row, account_col = cfg_cell(config, "account_cell", DEFAULT_CONFIG["account_cell"])
    account_raw = str(ws.cell(row=account_row, column=account_col).value or "").strip()
    parts = account_raw.split(
        cfg_str(
            config,
            "account_currency_separator",
            DEFAULT_CONFIG["account_currency_separator"],
        )
    )
    account_number = parts[0].strip() if parts else ""
    currency = (
        parts[1].strip()
        if len(parts) > 1
        else cfg_str(config, "default_currency", DEFAULT_CONFIG["default_currency"])
    )

    start_row, start_col = cfg_cell(
        config, "period_start_cell", DEFAULT_CONFIG["period_start_cell"]
    )
    end_row, end_col = cfg_cell(
        config, "period_end_cell", DEFAULT_CONFIG["period_end_cell"]
    )
    date_formats = DEFAULT_CONFIG["date_formats"]
    period_start = parse_date_str(
        str(ws.cell(row=start_row, column=start_col).value or ""),
        config,
        date_formats,
    )
    period_end = parse_date_str(
        str(ws.cell(row=end_row, column=end_col).value or ""),
        config,
        date_formats,
    )

    transaction_count = 0
    data_start_row = cfg_int(config, "data_start_row", DEFAULT_CONFIG["data_start_row"])
    max_columns = cfg_int(config, "max_columns", DEFAULT_CONFIG["max_columns"])
    description_column = cfg_int(
        config, "description_column", DEFAULT_CONFIG["description_column"]
    ) - 1
    for row in ws.iter_rows(min_row=data_start_row, max_col=max_columns):
        if row[description_column].value is not None:
            transaction_count += 1

    wb.close()

    if not period_start or not period_end:
        logger.warning(f"Could not parse date range from {xlsx_path.name}")
        return None

    logger.debug(
        f"[BANK-PARSE] {xlsx_path.name}: Millennium BCP, "
        f"account={account_number}, {period_start} to {period_end}, "
        f"{transaction_count} transactions"
    )

    return BankStatementData(
        bank_format=BankFormat.MILLENNIUM_BCP,
        account_number=account_number,
        currency=currency,
        period_start=period_start,
        period_end=period_end,
        transaction_count=transaction_count,
        issuing_party=cfg_str(config, "issuer_party", DEFAULT_CONFIG["issuer_party"]),
        issuing_party_raw=cfg_str(
            config, "issuer_party_raw", DEFAULT_CONFIG["issuer_party_raw"]
        ),
    )


def load_transactions(
    xlsx_path: Path,
    config: dict[str, object] | None = None,
) -> list[BankTransactionRecord] | None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    if not can_parse(ws, config=config):
        wb.close()
        return None

    transactions = []
    data_start_row = cfg_int(config, "data_start_row", DEFAULT_CONFIG["data_start_row"])
    max_columns = cfg_int(config, "max_columns", DEFAULT_CONFIG["max_columns"])
    description_column = cfg_int(
        config, "description_column", DEFAULT_CONFIG["description_column"]
    ) - 1
    amount_column = cfg_int(config, "amount_column", DEFAULT_CONFIG["amount_column"]) - 1
    currency_column = cfg_int(
        config, "currency_column", DEFAULT_CONFIG["currency_column"]
    ) - 1
    notes_column = cfg_int(config, "notes_column", DEFAULT_CONFIG["notes_column"]) - 1
    treated_column = cfg_int(config, "treated_column", DEFAULT_CONFIG["treated_column"]) - 1
    untreated_values = {
        value.lower()
        for value in cfg_sequence(
            config, "untreated_values", DEFAULT_CONFIG["untreated_values"]
        )
    }
    default_currency = cfg_str(config, "default_currency", DEFAULT_CONFIG["default_currency"])
    date_formats = DEFAULT_CONFIG["date_formats"]
    for row in ws.iter_rows(min_row=data_start_row, max_col=max_columns):
        if row[description_column].value is None:
            continue

        treated_val = str(row[treated_column].value or "").strip()
        if treated_val.lower() not in untreated_values:
            continue

        amount = parse_bank_amount(row[amount_column].value)
        if amount is None:
            continue

        transactions.append({
            "row_number": row[0].row,
            "date_posting": parse_date_cell(row[0].value, config, date_formats),
            "date_value": parse_date_cell(row[1].value, config, date_formats),
            "description": str(row[description_column].value or "").strip(),
            "amount": amount,
            "currency": str(row[currency_column].value or default_currency).strip(),
            "notes": str(row[notes_column].value or "").strip(),
            "treated": treated_val,
        })

    wb.close()
    return transactions
