"""BPI bank statement parser."""

import re
from pathlib import Path

import openpyxl

from papertrail.bank_statement.models import (
    BankFormat,
    BankStatementData,
    BankTransactionRecord,
    parse_bank_amount,
)
from papertrail.bank_statement.parser_utils import (
    cfg_cell,
    cfg_int,
    cfg_str,
    expected_headers,
    normalized_headers,
    parse_date_cell,
    parse_date_str,
)
from papertrail.logging_utils import get_logger

logger = get_logger("bank_statement")

FORMAT = BankFormat.BPI

DEFAULT_CONFIG: dict[str, object] = {
    "header_row": 18,
    "data_start_row": 19,
    "scan_columns": 7,
    "expected_headers": ("data mov.", "descricao do movimento", "valor em eur"),
    "date_formats": ("%d-%m-%Y", "%d/%m/%Y"),
    "account_cell": (7, 3),
    "account_currency_pattern": r"([\d\-.]+)\s*\((\w+)\)",
    "default_currency": "EUR",
    "issuer_party": "BPI",
    "issuer_party_raw": "BPI",
    "max_columns": 4,
    "description_column": 3,
    "amount_column": 4,
}


def can_parse(ws, config: dict[str, object] | None = None) -> bool:
    """Detect BPI format by checking column headers in row 18."""
    header_row = cfg_int(config, "header_row", DEFAULT_CONFIG["header_row"])
    scan_columns = cfg_int(config, "scan_columns", DEFAULT_CONFIG["scan_columns"])
    return expected_headers(config, DEFAULT_CONFIG["expected_headers"]).issubset(
        normalized_headers(ws, row=header_row, columns=scan_columns)
    )


def parse(xlsx_path: Path, config: dict[str, object] | None = None) -> BankStatementData | None:
    """Parse a BPI bank statement XLSX."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    if not can_parse(ws, config=config):
        wb.close()
        return None

    account_row, account_col = cfg_cell(config, "account_cell", DEFAULT_CONFIG["account_cell"])
    account_raw = str(ws.cell(row=account_row, column=account_col).value or "").strip()
    account_pattern = cfg_str(
        config,
        "account_currency_pattern",
        DEFAULT_CONFIG["account_currency_pattern"],
    )
    m = re.match(account_pattern, account_raw)
    if m:
        account_number = m.group(1).strip()
        currency = m.group(2).strip()
    else:
        account_number = account_raw
        currency = cfg_str(config, "default_currency", DEFAULT_CONFIG["default_currency"])

    dates = []
    transaction_count = 0
    data_start_row = cfg_int(config, "data_start_row", DEFAULT_CONFIG["data_start_row"])
    max_columns = cfg_int(config, "max_columns", DEFAULT_CONFIG["max_columns"])
    date_formats = DEFAULT_CONFIG["date_formats"]
    for row in ws.iter_rows(min_row=data_start_row, max_col=max_columns):
        cell_val = row[0].value
        if cell_val is None:
            continue
        date_str = str(cell_val).strip()
        parsed = parse_date_str(date_str, config, date_formats)
        if parsed:
            dates.append(parsed)
            transaction_count += 1

    wb.close()

    if not dates:
        logger.warning(f"No transaction dates found in {xlsx_path.name}")
        return None

    period_start = min(dates)
    period_end = max(dates)

    logger.debug(
        f"[BANK-PARSE] {xlsx_path.name}: BPI, "
        f"account={account_number}, {period_start} to {period_end}, "
        f"{transaction_count} transactions"
    )

    return BankStatementData(
        bank_format=BankFormat.BPI,
        account_number=account_number,
        currency=currency,
        period_start=period_start,
        period_end=period_end,
        transaction_count=transaction_count,
        issuing_party=cfg_str(config, "issuer_party", DEFAULT_CONFIG["issuer_party"]),
        issuing_party_raw=cfg_str(config, "issuer_party_raw", DEFAULT_CONFIG["issuer_party_raw"]),
    )


def load_transactions(
    xlsx_path: Path,
    config: dict[str, object] | None = None,
) -> list[BankTransactionRecord] | None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    if not can_parse(ws, config=config):
        wb.close()
        return None

    transactions = []
    data_start_row = cfg_int(config, "data_start_row", DEFAULT_CONFIG["data_start_row"])
    max_columns = cfg_int(config, "max_columns", DEFAULT_CONFIG["max_columns"])
    amount_column = cfg_int(config, "amount_column", DEFAULT_CONFIG["amount_column"]) - 1
    description_column = cfg_int(
        config, "description_column", DEFAULT_CONFIG["description_column"]
    ) - 1
    default_currency = cfg_str(config, "default_currency", DEFAULT_CONFIG["default_currency"])
    date_formats = DEFAULT_CONFIG["date_formats"]
    for row in ws.iter_rows(min_row=data_start_row, max_col=max_columns):
        if row[0].value is None:
            continue

        amount = parse_bank_amount(row[amount_column].value)
        if amount is None:
            continue

        transactions.append({
            "row_number": row[0].row,
            "date_posting": parse_date_cell(row[0].value, config, date_formats),
            "date_value": parse_date_cell(row[1].value, config, date_formats),
            "description": str(row[description_column].value or "").strip(),
            "amount": amount,
            "currency": default_currency,
            "notes": "",
            "treated": "",
        })

    wb.close()
    return transactions
