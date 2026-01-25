"""
papertrail - AI-powered PDF document classification and organization.

Main CLI entry point for processing PDF documents.
"""

import os
import re
import json
import shutil
import hashlib
import unicodedata
import argparse
import subprocess
import sys
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd
from tqdm import tqdm

# Import from papertrail package
from papertrail.config import (
    load_env,
    get_config_paths,
    get_openai_client,
    set_current_profile,
    get_current_profile,
)
from papertrail.profiles import (
    load_profile,
    list_available_profiles,
    get_profiles_dir,
    ProfileNotFoundError,
    ProfileError,
)
from papertrail.hashing import hash_file_fast, hash_file_content, HashCache
from papertrail.logging_utils import setup_failure_logger, log_failure, setup_logging, get_logger
from papertrail.models import (
    DocumentMetadata,
    DocumentMetadataRaw,
    normalize_enum_field_in_dict,
)
from papertrail.enums import reset_enum_cache
from papertrail.llm import (
    get_system_prompt_raw_extraction,
    TOOLS_RAW_EXTRACTION,
    normalize_metadata,
)
from papertrail.mappings import MappingsManager
from papertrail.pdf import render_pdf_to_images, find_pdf_files, get_page_count
from papertrail.metadata import (
    build_hash_index,
    get_unique_dates,
    save_metadata_json,
    iter_json_files,
)

# ------------------- LOGGING -------------------

# Module-level logger (configured by setup_logging in main())
logger = get_logger('cli')

# ------------------- CONFIG -------------------

from dataclasses import dataclass

@dataclass
class AppContext:
    """Runtime application context holding all initialized resources."""
    config_paths: dict
    model_id: str
    openai_client: any
    mappings_manager: any

# Single global context instance
_ctx: AppContext | None = None


def get_ctx() -> AppContext:
    """Get the application context. Raises if not initialized."""
    if _ctx is None:
        raise RuntimeError("Application context not initialized. Call initialize_config() first.")
    return _ctx


def initialize_config(profile_name: Optional[str] = None) -> None:
    """
    Initialize configuration from profile.

    Args:
        profile_name: Profile name to load, or None for auto-detection (uses 'default')

    Raises:
        ProfileNotFoundError: If specified profile doesn't exist
        ProfileError: If profile loading fails
    """
    global _ctx

    load_env()

    profiles_dir = get_profiles_dir()
    profiles_exist = profiles_dir.exists() and any(profiles_dir.glob("*.yaml"))

    if not profiles_exist:
        raise ProfileNotFoundError(
            "No profiles found. Create a profile from profiles/*.yaml.example. "
            "See profiles/README.md for documentation."
        )

    if profile_name is not None:
        profile = load_profile(profile_name)
        logger.info(f"Using profile: {profile.profile.name}")
    else:
        available = list_available_profiles()
        if "default" in available:
            profile = load_profile("default")
            logger.info(f"Using profile: {profile.profile.name} (auto-detected)")
        else:
            raise ProfileNotFoundError(
                f"No 'default' profile found. Available profiles: {', '.join(available)}. "
                "Use --profile to specify one, or create profiles/default.yaml."
            )

    if profile.profile.description:
        logger.info(f"  {profile.profile.description}")

    set_current_profile(profile)
    reset_enum_cache()

    config_dir = Path(__file__).parent / "config"
    mappings_path = config_dir / "mappings.yaml"

    _ctx = AppContext(
        config_paths=get_config_paths(),
        model_id=profile.openrouter.model_id,
        openai_client=get_openai_client(),
        mappings_manager=MappingsManager(mappings_path),
    )

# ------------------- UTILS -------------------

def sanitize_filename_component(s: str) -> str:
    """Sanitize a string for use in a filename."""
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    s = re.sub(r'[\\/*?:"<>|()\[\],]', '', s).strip()
    return re.sub(r'\s+', ' ', s)


def file_name_from_metadata(metadata: DocumentMetadata, file_hash: str) -> str:
    """Generate a filename from metadata."""
    parts = [
        sanitize_filename_component(metadata.issue_date),
        sanitize_filename_component(metadata.document_type.value),
        sanitize_filename_component(metadata.issuing_party.value)
    ]

    if metadata.service_name:
        parts.append(sanitize_filename_component(metadata.service_name))

    if metadata.total_amount is not None:
        amount = f"{metadata.total_amount:.0f}" if metadata.total_amount.is_integer() else f"{metadata.total_amount:.2f}"
        currency = metadata.total_amount_currency or ""
        parts.append(sanitize_filename_component(f"{amount} {currency}".strip()))

    parts.append(f"{file_hash}.pdf")
    return " - ".join(parts).lower()


# ------------------- CLASSIFICATION -------------------

def classify_pdf_document(pdf_path: Path, file_hash: str, failure_logger=None) -> DocumentMetadata:
    """Classify a PDF document using the LLM."""
    ctx = get_ctx()

    try:
        images_b64 = render_pdf_to_images(pdf_path)
    except Exception as e:
        log_failure(failure_logger, pdf_path, e)
        raise RuntimeError(f"Failed to render PDF image: {pdf_path}") from e

    try:
        user_content = [
            {
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}
            }
            for img_b64 in images_b64
        ]

        messages = [
            {"role": "system", "content": get_system_prompt_raw_extraction()},
            {"role": "user", "content": user_content},
        ]

        response = ctx.openai_client.chat.completions.create(
            model=ctx.model_id,
            max_tokens=4096,
            temperature=0,
            messages=messages,
            tools=TOOLS_RAW_EXTRACTION,
            tool_choice={"type": "function", "function": {"name": "extract_document_metadata"}},
        )

        tool_calls = response.choices[0].message.tool_calls
        if not tool_calls:
            raise ValueError("OpenRouter did not return structured classification.")

        args = tool_calls[0].function.arguments
        raw_metadata = DocumentMetadataRaw.model_validate_json(args)

        normalized_doc_type, normalized_issuing_party = normalize_metadata(
            raw_metadata, ctx.openai_client, ctx.model_id, mappings=ctx.mappings_manager
        )

        metadata = DocumentMetadata(
            issue_date=raw_metadata.issue_date,
            document_type=normalized_doc_type,
            issuing_party=normalized_issuing_party,
            service_name=raw_metadata.service_name,
            total_amount=raw_metadata.total_amount,
            total_amount_currency=raw_metadata.total_amount_currency,
            confidence=raw_metadata.confidence,
            reasoning=raw_metadata.reasoning,
            content_hash=file_hash,
            document_type_raw=raw_metadata.document_type,
            issuing_party_raw=raw_metadata.issuing_party,
        )

        now = datetime.now().strftime("%Y-%m-%d")
        metadata.create_date = now
        metadata.update_date = now
        return metadata
    except Exception as e:
        log_failure(failure_logger, pdf_path, e)
        raise RuntimeError(f"Classification failed for: {pdf_path}") from e


# ------------------- RENAMING & PROCESSING -------------------

def rename_single_pdf(pdf_path: Path, content_hash: str, processed_path: Path, known_hashes: set, failure_logger=None):
    """Process and rename a single PDF file."""
    try:
        file_hash = hash_file_fast(pdf_path)
        metadata = classify_pdf_document(pdf_path, content_hash, failure_logger)
        metadata.file_hash = file_hash

        filename = file_name_from_metadata(metadata, content_hash)
        new_pdf_path = processed_path / filename

        shutil.copy2(pdf_path, new_pdf_path)
        save_metadata_json(new_pdf_path, metadata)

        known_hashes.add(content_hash)
        known_hashes.add(file_hash)
        logger.info(f"Processed: {pdf_path.name} -> {filename}")
    except Exception as e:
        log_failure(failure_logger, pdf_path, e)
        logger.error(f"Failed to process {pdf_path.name}: {e}")


def rename_pdf_files(pdf_paths, file_hash_map, known_hashes, processed_path, failure_logger=None):
    """Rename multiple PDF files."""
    for pdf_path in tqdm(pdf_paths):
        rename_single_pdf(pdf_path, file_hash_map[pdf_path], processed_path, known_hashes, failure_logger)


def validate_metadata(output_path: Path):
    """Validate metadata files and their corresponding PDFs.

    Uses caching and parallelization to speed up content hash computation:
    1. Fast file hash is computed (cheap, ~0.05s)
    2. Cache is checked for existing file_hash -> content_hash mapping
    3. Cache misses are computed in parallel using ProcessPoolExecutor
    """
    from concurrent.futures import ProcessPoolExecutor, as_completed

    valid_entries = []
    errors = []
    json_files = list(output_path.rglob("*.json"))

    # Initialize hash cache
    cache = HashCache()
    logger.info(f"Hash cache loaded with {len(cache)} entries")

    # Phase 1: Collect all PDF paths and their expected hashes
    pdf_info = []  # List of (metadata_path, pdf_path, expected_hash, metadata)
    for metadata_path in json_files:
        try:
            with open(metadata_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            metadata = DocumentMetadata.model_validate(data)

            content_hash = metadata.content_hash
            if not content_hash:
                errors.append((metadata_path, "Missing 'content_hash' in metadata."))
                continue

            pdf_path = metadata_path.with_suffix(".pdf")
            if not pdf_path.exists():
                errors.append((metadata_path, f"Missing PDF for metadata: {pdf_path.name}"))
                continue

            pdf_info.append((metadata_path, pdf_path, content_hash, metadata))

        except Exception as e:
            errors.append((metadata_path, str(e)))

    if not pdf_info:
        if errors:
            logger.warning("Validation errors found:")
            for meta_path, err in errors:
                logger.warning(f"- {meta_path}: {err}")
        return valid_entries

    # Phase 2: Compute fast hashes and check cache
    logger.info(f"Computing fast hashes for {len(pdf_info)} PDFs...")
    hash_results = {}  # pdf_path -> content_hash
    uncached = []  # List of (pdf_path, file_hash)

    for _, pdf_path, _, _ in tqdm(pdf_info, desc="Fast hashing"):
        file_hash = hash_file_fast(pdf_path)
        cached_content_hash = cache.get(file_hash)
        if cached_content_hash:
            hash_results[pdf_path] = cached_content_hash
        else:
            uncached.append((pdf_path, file_hash))

    cache_hits = len(pdf_info) - len(uncached)
    logger.info(f"  -> Cache hits: {cache_hits}, Cache misses: {len(uncached)}")

    # Phase 3: Parallel content hashing for uncached PDFs
    if uncached:
        logger.info(f"Computing content hashes for {len(uncached)} uncached PDFs...")
        with ProcessPoolExecutor() as executor:
            # Submit all jobs
            futures = {executor.submit(hash_file_content, pdf_path): (pdf_path, file_hash)
                       for pdf_path, file_hash in uncached}

            # Collect results with progress bar
            for future in tqdm(as_completed(futures), total=len(futures), desc="Content hashing"):
                pdf_path, file_hash = futures[future]
                try:
                    content_hash = future.result()
                    hash_results[pdf_path] = content_hash
                    cache.set(file_hash, content_hash)
                except Exception as e:
                    # Find the metadata_path for this pdf_path
                    for metadata_path, p, _, _ in pdf_info:
                        if p == pdf_path:
                            errors.append((metadata_path, f"Content hashing failed: {e}"))
                            break

        # Save cache with new entries
        cache.save()
        logger.info(f"Hash cache saved with {len(cache)} entries")

    # Phase 4: Validate using precomputed hashes
    for metadata_path, pdf_path, expected_hash, metadata in pdf_info:
        actual_hash = hash_results.get(pdf_path)
        if actual_hash is None:
            continue  # Error already recorded

        if expected_hash != actual_hash:
            errors.append((metadata_path, f"Hash mismatch: metadata content_hash is '{expected_hash}', actual is '{actual_hash}'."))
            continue

        if expected_hash not in pdf_path.name:
            errors.append((metadata_path, f"Filename '{pdf_path.name}' does not include the expected hash '{expected_hash}'."))
            continue

        valid_entries.append((pdf_path, metadata))

    if errors:
        logger.warning("Validation errors found:")
        for meta_path, err in errors:
            logger.warning(f"- {meta_path}: {err}")
    else:
        logger.info("All metadata files passed validation.")

    return valid_entries


def validate_merged_pdf(folder_path: Path) -> bool:
    """
    Validate that merged_all.pdf has the correct page count.

    Compares the page count of merged_all.pdf against the sum of
    page counts of all other PDFs in the folder.

    Returns True if valid, raises AssertionError if mismatch.
    """
    merged_path = folder_path / "merged_all.pdf"
    if not merged_path.exists():
        logger.info(f"  No merged_all.pdf found in {folder_path}")
        return True

    source_pdfs = [p for p in folder_path.glob("*.pdf") if p.name != "merged_all.pdf"]
    expected_pages = sum(get_page_count(pdf) for pdf in source_pdfs)

    actual_pages = get_page_count(merged_path)

    if actual_pages != expected_pages:
        raise AssertionError(
            f"Merged PDF page count mismatch in {folder_path}: "
            f"expected {expected_pages} pages (from {len(source_pdfs)} files), "
            f"got {actual_pages} pages"
        )

    logger.info(f"  Merge validation passed: {actual_pages} pages from {len(source_pdfs)} files")
    return True


def export_metadata_to_excel(processed_path: Path, excel_output_path: str):
    """Export metadata to an Excel file."""
    from enum import Enum
    from papertrail.metadata import load_json_files_parallel

    metadata_list = []

    for metadata_path, metadata in load_json_files_parallel(processed_path, validate=True, show_progress=True, progress_desc="Collecting metadata"):
        metadata_dict = metadata.model_dump()

        metadata_dict.pop("reasoning", None)

        pdf_path = metadata_path.with_suffix(".pdf")
        filename = pdf_path.name if pdf_path.exists() else ""
        metadata_dict["filename"] = filename
        metadata_dict["filename_length"] = len(filename)

        try:
            date_parts = metadata.issue_date.split('-')
            metadata_dict["year"] = int(date_parts[0])
            metadata_dict["month"] = int(date_parts[1])
        except (IndexError, ValueError, AttributeError):
            metadata_dict["year"] = None
            metadata_dict["month"] = None

        # Normalize enum fields
        normalize_enum_field_in_dict(metadata_dict, "document_type", "DocumentType")
        normalize_enum_field_in_dict(metadata_dict, "issuing_party", "IssuingParty")

        metadata_list.append(metadata_dict)

    if metadata_list:
        df = pd.DataFrame(metadata_list)
        ordered_cols = [
            "confidence", "issue_date", "year", "month", "content_hash", "file_hash",
            "filename", "filename_length", "document_type", "document_type_raw",
            "issuing_party", "issuing_party_raw", "service_name",
            "total_amount", "total_amount_currency"
        ]
        extra_cols = [col for col in df.columns if col not in ordered_cols]
        df = df[ordered_cols + extra_cols]

        if "issue_date" in df.columns:
            df = df.sort_values(by="issue_date", ascending=False)

        with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            worksheet.freeze_panes = 'A2'

            from openpyxl.utils import get_column_letter
            for col in ordered_cols:
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1
                    col_letter = get_column_letter(col_idx)
                    values_lens = [len(str(val)) for val in df[col].values if val is not None]
                    max_len = max(values_lens + [len(col)])
                    worksheet.column_dimensions[col_letter].width = min(max_len + 2, 102)

            hidden_cols = ["year", "month", "filename_length"]
            for col in hidden_cols:
                if col in df.columns:
                    col_letter = get_column_letter(df.columns.get_loc(col) + 1)
                    worksheet.column_dimensions[col_letter].hidden = True

        logger.info(f"Exported {len(df)} entries to {excel_output_path}")
    else:
        logger.info("No valid metadata found to export.")


def copy_matching_files(
    processed_path: Path,
    regex_pattern: str,
    dest_folder: Path,
    incremental: bool = False
) -> dict:
    """
    Copy files matching regex pattern to destination.

    Args:
        processed_path: Source directory
        regex_pattern: Pattern to match filenames
        dest_folder: Destination directory
        incremental: If True, skip files that already exist with same content

    Returns:
        Stats dict with 'copied', 'skipped', 'total' counts
    """
    dest_folder.mkdir(parents=True, exist_ok=True)
    pattern = re.compile(regex_pattern)
    stats = {'copied': 0, 'skipped': 0, 'total': 0}

    for file in processed_path.iterdir():
        if not file.is_file():
            continue
        if file.suffix.lower() not in [".pdf", ".json"]:
            continue
        if not pattern.search(file.name):
            continue

        stats['total'] += 1
        dest_file = dest_folder / file.name

        should_copy = True
        if incremental and dest_file.exists():
            if file.stat().st_size == dest_file.stat().st_size:
                src_hash = hash_file_fast(file)
                dst_hash = hash_file_fast(dest_file)
                if src_hash == dst_hash:
                    should_copy = False
                    stats['skipped'] += 1

        if should_copy:
            shutil.copy2(file, dest_file)
            stats['copied'] += 1

    return stats


def calculate_directory_hash(directory: Path) -> str:
    """Calculate a hash representing all PDF files in the directory."""
    pdf_files = sorted(directory.glob("*.pdf"))
    if not pdf_files:
        return ""

    combined = []
    for pdf_file in pdf_files:
        file_hash = hash_file_fast(pdf_file)
        combined.append(f"{pdf_file.name}:{file_hash}")

    combined_str = "\n".join(combined)
    return hashlib.sha256(combined_str.encode()).hexdigest()[:16]


def directory_has_changed(directory: Path) -> bool:
    """Check if directory contents have changed since last check."""
    hash_file_path = directory / ".directory_hash"
    current_hash = calculate_directory_hash(directory)

    if not current_hash:
        return False

    if not hash_file_path.exists():
        with open(hash_file_path, "w") as f:
            f.write(current_hash)
        return True

    with open(hash_file_path, "r") as f:
        stored_hash = f.read().strip()

    if current_hash != stored_hash:
        with open(hash_file_path, "w") as f:
            f.write(current_hash)
        return True

    return False


def check_files_exist(target_folder: Path, validation_schema_path: Path):
    """Validate files exist based on a schema."""
    with open(validation_schema_path, "r", encoding="utf-8") as f:
        checks = json.load(f)

    json_files = list(target_folder.glob("*.json"))
    file_data = []
    for json_path in json_files:
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            file_data.append((json_path, data))
        except Exception as e:
            logger.warning(f"Skipping {json_path.name}: {e}")

    check_results = []
    for idx, check in enumerate(checks):
        found = any(
            all(str(data.get(k, "")).strip() == str(v).strip() for k, v in check.items())
            for _, data in file_data
        )
        check_results.append((found, idx, check))

    all_passed = all(found for found, _, _ in check_results)

    # Sort: found items first, then by index
    sorted_results = sorted(check_results, key=lambda x: (not x[0], x[1]))
    for found, idx, check in sorted_results:
        status = "[OK]" if found else "[FAIL]"
        result = "FOUND" if found else "NOT FOUND"
        if found:
            logger.info(f"{status} {check} -- {result}")
        else:
            logger.warning(f"{status} {check} -- {result}")

    if all_passed:
        logger.info("All file existence checks passed.")
    else:
        logger.warning("Some file existence checks failed.")


# ------------------- TASK HANDLERS -------------------

def task_extract_new(processed_path: Path, raw_paths: list[Path]):
    """Extract and classify new PDF files."""
    log_path = processed_path / "classification_failures.log"
    failure_logger = setup_failure_logger(log_path)
    logger.debug(f"Logging failures to: {log_path}")

    logger.info("Building hash index from metadata files...")
    known_hashes = set(build_hash_index(processed_path).keys())

    logger.info("Scanning for new PDFs...")
    pdf_paths = find_pdf_files(raw_paths)
    logger.info(f"Found {len(pdf_paths)} PDFs in raw directories")

    logger.info("Stage 1: Quick filtering using fast file hashes...")
    fast_hash_map = {pdf: hash_file_fast(pdf) for pdf in tqdm(pdf_paths, desc="Fast hashing")}
    potentially_new = [pdf for pdf in pdf_paths if fast_hash_map[pdf] not in known_hashes]

    already_processed = len(pdf_paths) - len(potentially_new)
    logger.info(f"  -> Skipped {already_processed} already-processed files")
    logger.info(f"  -> {len(potentially_new)} files need content-based hashing")

    if not potentially_new:
        logger.info("No new PDFs to process.")
        return

    logger.info(f"Stage 2: Content-based hashing for {len(potentially_new)} new files...")
    content_hash_map = {}

    for pdf in tqdm(potentially_new, desc="Content hashing"):
        try:
            content_hash = hash_file_content(pdf)
            content_hash_map[pdf] = content_hash
        except Exception as e:
            logger.error(f"Error hashing {pdf.name}: {e}")

    files_to_process = [pdf for pdf in potentially_new if content_hash_map.get(pdf) not in known_hashes]
    logger.info(f"Found {len(files_to_process)} truly new PDFs to process.")

    if files_to_process:
        rename_pdf_files(files_to_process, content_hash_map, known_hashes, processed_path, failure_logger)

    logger.info("Extraction complete.")


def task_rename_files(processed_path: Path):
    """Rename existing PDF files based on metadata."""
    from papertrail.metadata import load_json_files_parallel

    logger.info("Renaming existing PDF files and metadata based on metadata...")

    valid_entries = []

    for metadata_path, metadata in load_json_files_parallel(processed_path, validate=True, show_progress=True, progress_desc="Validating metadata"):
        pdf_path = metadata_path.with_suffix(".pdf")
        if not pdf_path.exists():
            logger.warning(f"Skipping {metadata_path.name}: PDF file not found")
            continue

        valid_entries.append((pdf_path, metadata))

    logger.info(f"Found {len(valid_entries)} files to rename")

    renamed_count = 0
    for old_pdf_path, metadata in valid_entries:
        content_hash = metadata.content_hash
        new_filename = file_name_from_metadata(metadata, content_hash)
        new_pdf_path = processed_path / new_filename
        new_metadata_path = new_pdf_path.with_suffix(".json")

        if old_pdf_path == new_pdf_path:
            continue

        try:
            old_metadata_path = old_pdf_path.with_suffix(".json")
            shutil.move(old_pdf_path, new_pdf_path)
            shutil.move(old_metadata_path, new_metadata_path)
            renamed_count += 1
            if renamed_count <= 10 or renamed_count % 100 == 0:
                logger.info(f"[{renamed_count}] Renamed: {old_pdf_path.name} -> {new_filename}")
        except Exception as e:
            logger.error(f"Failed to rename {old_pdf_path.name}: {e}")

    logger.info(f"Renaming complete. Renamed {renamed_count} files.")


def task_export_all_dates(
    processed_path: Path,
    export_base_dir: Path,
    run_merge: bool = False,
):
    """Export files for all unique dates found in processed files."""
    processed_path = Path(processed_path)
    export_base_dir = Path(export_base_dir)

    logger.info("Scanning for unique dates in processed files...")
    all_dates = get_unique_dates(processed_path)

    if not all_dates:
        logger.info("No dates found in processed files.")
        return

    logger.info(f"Found {len(all_dates)} unique dates: {', '.join(all_dates[:10])}{' ...' if len(all_dates) > 10 else ''}")

    total_copied = 0
    total_skipped = 0
    changed_directories = []

    for date in all_dates:
        export_date_dir = export_base_dir / date
        logger.info(f"[{date}] Processing...")

        stats = copy_matching_files(processed_path, date, export_date_dir, incremental=True)
        total_copied += stats['copied']
        total_skipped += stats['skipped']

        if stats['total'] == 0:
            logger.info(f"  No files match date pattern '{date}'")
        else:
            logger.info(f"  Copied: {stats['copied']}, Skipped: {stats['skipped']}, Total: {stats['total']}")

        if stats['copied'] > 0:
            changed_directories.append(export_date_dir)
        elif stats['total'] > 0:
            if export_date_dir.exists() and directory_has_changed(export_date_dir):
                changed_directories.append(export_date_dir)

    logger.info("=== Summary ===")
    logger.info(f"Processed {len(all_dates)} date(s)")
    logger.info(f"Total files copied: {total_copied}")
    logger.info(f"Total files skipped (unchanged): {total_skipped}")
    logger.info(f"Directories with changes: {len(changed_directories)}")

    if run_merge and changed_directories:
        logger.info("=== Running PDF Merge ===")
        from shutil import which
        if which("pdf-merger") is None:
            logger.warning("pdf-merger tool not found in PATH. Skipping merge step.")
        else:
            for export_dir in changed_directories:
                logger.info(f"Merging PDFs in {export_dir}...")
                result = subprocess.run(f'pdf-merger "{export_dir}"', shell=True, text=True)
                if result.returncode == 0:
                    logger.info("  Merge completed successfully")
                    validate_merged_pdf(export_dir)
                else:
                    logger.error(f"  Merge failed with exit code {result.returncode}")

    logger.info("Export all dates complete.")


def task_bootstrap_mappings(processed_path: Path, mappings_mgr):
    """Populate mappings from existing metadata JSON files."""
    if mappings_mgr is None:
        logger.error("Mappings manager not initialized.")
        return

    json_files = list(processed_path.rglob("*.json"))
    if not json_files:
        logger.info(f"No metadata files found in {processed_path}")
        return

    doc_type_count = 0
    issuer_count = 0
    skipped = 0

    for metadata_path in tqdm(json_files, desc="Scanning metadata"):
        try:
            with open(metadata_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            doc_type_raw = data.get("document_type_raw")
            doc_type = data.get("document_type")
            issuing_party_raw = data.get("issuing_party_raw")
            issuing_party = data.get("issuing_party")

            if doc_type_raw and doc_type and doc_type != "$UNKNOWN$":
                existing = mappings_mgr.get_mapping(doc_type_raw, "document_types")
                if existing is None:
                    mappings_mgr.add_mapping(
                        doc_type_raw, doc_type, "document_types", confirmed=True, save=False
                    )
                    doc_type_count += 1

            if issuing_party_raw and issuing_party and issuing_party != "$UNKNOWN$":
                existing = mappings_mgr.get_mapping(issuing_party_raw, "issuing_parties")
                if existing is None:
                    mappings_mgr.add_mapping(
                        issuing_party_raw, issuing_party, "issuing_parties", confirmed=True, save=False
                    )
                    issuer_count += 1

        except Exception as e:
            skipped += 1
            if skipped <= 5:
                logger.warning(f"Skipping {metadata_path.name}: {e}")

    mappings_mgr._save()

    logger.info("Bootstrap complete:")
    logger.info(f"  Document type mappings added: {doc_type_count}")
    logger.info(f"  Issuing party mappings added: {issuer_count}")
    logger.info(f"  Files skipped: {skipped}")
    logger.info(f"Mappings saved to: {mappings_mgr.path}")

    stats = mappings_mgr.get_stats()
    logger.info("Current mappings stats:")
    for field, counts in stats.items():
        logger.info(f"  {field}: {counts['confirmed']} confirmed, {counts['auto']} auto, {counts['canonicals']} canonicals")


def task_review_mappings(mappings_mgr):
    """Interactive review of auto-added mappings."""
    if mappings_mgr is None:
        logger.error("Mappings manager not initialized.")
        return

    doc_auto = mappings_mgr.get_auto_mappings("document_types")
    issuer_auto = mappings_mgr.get_auto_mappings("issuing_parties")

    total_pending = len(doc_auto) + len(issuer_auto)

    if total_pending == 0:
        logger.info("No auto-added mappings pending review.")
        stats = mappings_mgr.get_stats()
        logger.info("Current mappings stats:")
        for field, counts in stats.items():
            logger.info(f"  {field}: {counts['confirmed']} confirmed, {counts['auto']} auto")
        return

    print("=" * 60)
    print("AUTO-ADDED MAPPINGS AWAITING REVIEW")
    print("=" * 60)
    print()

    if doc_auto:
        print(f"Document Types ({len(doc_auto)} pending):")
        for i, (raw, canonical) in enumerate(doc_auto.items(), 1):
            print(f"  {i}. \"{raw}\" -> \"{canonical}\"")
        print()

    if issuer_auto:
        print(f"Issuing Parties ({len(issuer_auto)} pending):")
        for i, (raw, canonical) in enumerate(issuer_auto.items(), 1):
            print(f"  {i}. \"{raw}\" -> \"{canonical}\"")
        print()

    print("Options:")
    print("  [a] Confirm ALL mappings")
    print("  [r] Review one-by-one")
    print("  [q] Quit without changes")
    print()

    choice = input("Select option: ").strip().lower()

    if choice == 'a':
        doc_confirmed = mappings_mgr.confirm_all("document_types", save=False)
        issuer_confirmed = mappings_mgr.confirm_all("issuing_parties", save=True)
        logger.info(f"Confirmed {doc_confirmed} document type mappings and {issuer_confirmed} issuer mappings.")

    elif choice == 'r':
        _review_field_mappings(mappings_mgr, "document_types", doc_auto)
        _review_field_mappings(mappings_mgr, "issuing_parties", issuer_auto)
        logger.info("Review complete.")

    else:
        logger.info("No changes made.")


def _review_field_mappings(mappings_mgr, field: str, mappings_dict: dict):
    """Helper to review mappings for a single field."""
    if not mappings_dict:
        return

    field_label = "Document Type" if field == "document_types" else "Issuing Party"
    print(f"\n--- Reviewing {field_label} Mappings ---")

    for raw, canonical in list(mappings_dict.items()):
        print(f"\n\"{raw}\" -> \"{canonical}\"")
        print("  [c] Confirm  [e] Edit canonical  [r] Reject  [s] Skip")
        action = input("  Action: ").strip().lower()

        if action == 'c':
            mappings_mgr.confirm_mapping(raw, field, save=True)
            print("  Confirmed.")
        elif action == 'e':
            new_canonical = input("  Enter new canonical value: ").strip()
            if new_canonical:
                mappings_mgr.update_mapping(raw, new_canonical, field, confirm=True, save=True)
                print(f"  Updated to \"{new_canonical}\" and confirmed.")
            else:
                print("  No change (empty input).")
        elif action == 'r':
            mappings_mgr.reject_mapping(raw, field, save=True)
            print("  Rejected (removed).")
        else:
            print("  Skipped.")


def task_add_canonical(mappings_mgr, field: str, canonical: str):
    """Add a new canonical value to the mappings."""
    if mappings_mgr is None:
        logger.error("Mappings manager not initialized.")
        return

    field_map = {
        "document_type": "document_types",
        "document_types": "document_types",
        "issuing_party": "issuing_parties",
        "issuing_parties": "issuing_parties",
    }

    normalized_field = field_map.get(field.lower())
    if not normalized_field:
        logger.error(f"Unknown field '{field}'. Use 'document_type' or 'issuing_party'.")
        return

    if mappings_mgr.add_canonical(normalized_field, canonical):
        logger.info(f"Added canonical '{canonical}' to {normalized_field}.")
        logger.info(f"Current canonicals: {', '.join(mappings_mgr.get_canonicals(normalized_field))}")
    else:
        logger.info(f"Canonical '{canonical}' already exists in {normalized_field}.")


def task_gmail_download():
    """Download email attachments from Gmail."""
    from datetime import timedelta
    from papertrail.gmail import download_gmail_attachments

    profile = get_current_profile()
    if not profile:
        logger.error("No profile is active.")
        sys.exit(1)

    raw_paths = profile.paths.raw
    processed_path_str = profile.paths.processed

    if not raw_paths or not processed_path_str:
        missing = []
        if not raw_paths:
            missing.append("paths.raw")
        if not processed_path_str:
            missing.append("paths.processed")
        logger.error(f"Missing required profile settings: {', '.join(missing)}")
        sys.exit(1)

    raw_path = Path(raw_paths[0])
    processed_path = Path(processed_path_str)

    raw_path.mkdir(parents=True, exist_ok=True)

    end_date = datetime.now()

    unique_dates = get_unique_dates(processed_path) if processed_path.exists() else []

    if unique_dates:
        most_recent = unique_dates[0]
        start_date = datetime.strptime(f"{most_recent}-01", "%Y-%m-%d")
        logger.info(f"Date range: {start_date.date()} to {end_date.date()}")
    else:
        start_date = end_date - timedelta(days=30)
        logger.info("No processed files found. Using default range: last 30 days")

    logger.info(f"Downloading attachments to: {raw_path}")

    stats = download_gmail_attachments(
        output_dir=raw_path,
        start_date=start_date,
        end_date=end_date,
    )

    logger.info("=== Gmail Download Summary ===")
    logger.info(f"Messages found: {stats['messages_found']}")
    logger.info(f"Messages processed: {stats['messages_processed']}")
    logger.info(f"Messages skipped: {stats['messages_skipped']}")
    logger.info(f"Attachments downloaded: {stats['attachments_downloaded']}")
    logger.info(f"Attachments failed: {stats['attachments_failed']}")
    logger.info(f"Bytes downloaded: {stats['bytes_downloaded']:,}")


# ------------------- PIPELINE -------------------

def run_step(cmd, step_desc):
    """Run a pipeline step."""
    logger.info(f"### {step_desc}...")
    result = subprocess.run(cmd, shell=True, text=True)
    if result.returncode != 0:
        logger.error(f"{step_desc} failed with exit code {result.returncode}.")
        sys.exit(result.returncode)
    logger.info(f"### {step_desc}... Finished.")


def pipeline(export_date_arg=None):
    """Run the full document processing pipeline."""
    from shutil import which
    from datetime import timedelta

    RAW_FILES_DIR = os.getenv("RAW_FILES_DIR")
    PROCESSED_FILES_DIR = os.getenv("PROCESSED_FILES_DIR")
    EXPORT_FILES_DIR = os.getenv("EXPORT_FILES_DIR")

    required_vars = [
        ("RAW_FILES_DIR", RAW_FILES_DIR),
        ("PROCESSED_FILES_DIR", PROCESSED_FILES_DIR),
        ("EXPORT_FILES_DIR", EXPORT_FILES_DIR),
    ]
    missing = [name for name, val in required_vars if not val]
    if missing:
        logger.error(f"Missing required .env variables: {', '.join(missing)}")
        sys.exit(1)

    for tool in ["mbox-extractor", "archive-extractor", "pdf-merger"]:
        if which(tool) is None:
            logger.error(f"Required tool '{tool}' not found in PATH. Please install it and try again.")
            sys.exit(1)

    if export_date_arg:
        export_date = export_date_arg
    else:
        today = datetime.now()
        first_of_this_month = today.replace(day=1)
        last_month = first_of_this_month - timedelta(days=1)
        export_date = last_month.strftime("%Y-%m")

    if not re.match(r"^\d{4}-\d{2}$", export_date):
        logger.error("The export_date must be in YYYY-MM format.")
        sys.exit(1)

    export_date_dir = os.path.join(EXPORT_FILES_DIR, export_date)

    # Get passwords from profile (inline or file reference)
    from papertrail.config import get_passwords, get_validations
    passwords, passwords_file = get_passwords()

    # Track temp files for cleanup
    temp_passwords_file = None
    temp_validations_file = None

    # Setup passwords file path for archive-extractor tool
    if passwords:
        if passwords_file:
            # Use profile file reference directly
            zip_passwords_file_path = passwords_file
            logger.debug(f"Using passwords file from profile: {zip_passwords_file_path}")
        else:
            # Create temp file for inline data
            temp_passwords = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.txt')
            temp_passwords.write('\n'.join(passwords))
            temp_passwords.close()
            zip_passwords_file_path = temp_passwords.name
            temp_passwords_file = temp_passwords.name
            logger.debug(f"Created temporary passwords file: {zip_passwords_file_path}")
    else:
        logger.warning("No passwords configured. Skipping password-protected archives.")
        zip_passwords_file_path = None

    # Get validations from profile (inline or file reference)
    validations, validations_file = get_validations()

    # Setup validations file path for check_files_exist task
    if validations and validations.get('rules'):
        if validations_file:
            # Use profile file reference directly
            validations_file_path = validations_file
            logger.debug(f"Using validations file from profile: {validations_file_path}")
        else:
            # Create temp file for inline data
            temp_validations = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json')
            json.dump(validations['rules'], temp_validations, indent=2)
            temp_validations.close()
            validations_file_path = temp_validations.name
            temp_validations_file = temp_validations.name
            logger.debug(f"Created temporary validations file: {validations_file_path}")
    else:
        validations_file_path = None

    processed_files_excel_path = Path(PROCESSED_FILES_DIR) / "processed_files.xlsx"
    raw_dirs = [p for p in RAW_FILES_DIR.split(';') if p]

    run_step(f'"{sys.executable}" "{__file__}" gmail_download', "Step 1: Download Gmail attachments")

    for rd in raw_dirs:
        run_step(f'mbox-extractor "{rd}"', "Step 2: Google Takeout mbox extraction")
        if zip_passwords_file_path:
            run_step(f'archive-extractor "{rd}" --passwords "{zip_passwords_file_path}"', "Step 3: Google Takeout zip extraction")
        else:
            run_step(f'archive-extractor "{rd}"', "Step 3: Google Takeout zip extraction")

    raw_dirs_arg = ";".join(raw_dirs)
    run_step(f'"{sys.executable}" "{__file__}" extract_new "{PROCESSED_FILES_DIR}" --raw_path "{raw_dirs_arg}"', "Step 4: Extract new documents")
    run_step(f'"{sys.executable}" "{__file__}" rename_files "{PROCESSED_FILES_DIR}"', "Step 5: Rename files and metadata")
    run_step(f'"{sys.executable}" "{__file__}" export_excel "{PROCESSED_FILES_DIR}" --excel_output_path "{processed_files_excel_path}"', "Step 6: Export metadata to Excel")
    run_step(f'"{sys.executable}" "{__file__}" copy_matching "{PROCESSED_FILES_DIR}" --regex_pattern "{export_date}" --copy_dest_folder "{export_date_dir}"', "Step 7: Copy matching documents")
    run_step(f'pdf-merger "{export_date_dir}"', "Step 8: Merge PDFs")
    validate_merged_pdf(Path(export_date_dir))
    if validations_file_path:
        run_step(f'"{sys.executable}" "{__file__}" check_files_exist "{export_date_dir}" --check_schema_path "{validations_file_path}"', "Step 9: Validate exported files")
    else:
        logger.info("Step 9: Skipping file validation (no validation rules configured in profile)")

    # Cleanup temporary files
    if temp_passwords_file:
        try:
            os.unlink(temp_passwords_file)
            logger.debug(f"Cleaned up temporary passwords file: {temp_passwords_file}")
        except Exception as e:
            logger.warning(f"Failed to cleanup temporary passwords file: {e}")

    if temp_validations_file:
        try:
            os.unlink(temp_validations_file)
            logger.debug(f"Cleaned up temporary validations file: {temp_validations_file}")
        except Exception as e:
            logger.warning(f"Failed to cleanup temporary validations file: {e}")

    logger.info("All steps completed successfully.")


def process_folder(task: str, processed_path: str, raw_paths=None, excel_output_path: str = None,
                   regex_pattern: str = None, copy_dest_folder: str = None, check_schema_path: str = None,
                   export_base_dir: str = None, run_merge: bool = False):
    """Dispatch to appropriate task handler."""
    if raw_paths is not None:
        raw_paths = [Path(p) for p in raw_paths]
    processed_path = Path(processed_path)
    processed_path.mkdir(parents=True, exist_ok=True)

    if task == "extract_new":
        task_extract_new(processed_path, raw_paths)
    elif task == "rename_files":
        task_rename_files(processed_path)
    elif task == "validate_metadata":
        logger.info("Validating existing metadata and PDFs...")
        validate_metadata(processed_path)
        logger.info("Validation complete.")
    elif task == "export_excel":
        logger.info("Exporting metadata to Excel...")
        export_metadata_to_excel(processed_path, excel_output_path)
        logger.info("Excel export complete.")
    elif task == "copy_matching":
        if not regex_pattern or not copy_dest_folder:
            logger.error("For 'copy_matching', --regex_pattern and --copy_dest_folder are required.")
            return
        stats = copy_matching_files(processed_path, regex_pattern, Path(copy_dest_folder))
        logger.info(f"Copied {stats['copied']} files matching '{regex_pattern}' to {copy_dest_folder}")
    elif task == "export_all_dates":
        task_export_all_dates(processed_path, Path(export_base_dir), run_merge)
    elif task == "check_files_exist":
        if not check_schema_path:
            logger.error("For 'check_files_exist', --check_schema_path is required.")
            return
        check_files_exist(processed_path, Path(check_schema_path))
        logger.info("File existence check complete.")
    else:
        logger.error("Invalid task specified.")


# ------------------- MAIN -------------------

def main():
    """Main CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Process a folder of PDF files.",
        epilog="Use --profile to select a configuration profile. "
               "Available profiles are listed in profiles/ directory."
    )
    parser.add_argument(
        "--profile",
        type=str,
        help="Configuration profile to use (e.g., 'default', 'personal', 'work'). "
             "If not specified, uses 'default' profile if available, otherwise legacy .env configuration."
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose output with timestamps and debug messages."
    )
    parser.add_argument("task", type=str, choices=[
        'extract_new', 'rename_files', 'validate_metadata', 'export_excel',
        'copy_matching', 'export_all_dates', 'check_files_exist', 'pipeline',
        'gmail_download', 'bootstrap_mappings', 'review_mappings', 'add_canonical'
    ], help="Task to perform.")
    parser.add_argument("processed_path", type=str, nargs='?', help="Path to output folder.")
    parser.add_argument("--raw_path", type=str, help="Path to documents folder(s). Use ';' to separate multiple paths.")
    parser.add_argument("--excel_output_path", type=str, help="Path to output Excel file.")
    parser.add_argument("--regex_pattern", type=str, help="Regex pattern for matching filenames.")
    parser.add_argument("--copy_dest_folder", type=str, help="Destination folder for copied files.")
    parser.add_argument("--export_base_dir", type=str, help="Base export directory.")
    parser.add_argument("--run_merge", action="store_true", help="Run PDF merge for changed directories.")
    parser.add_argument("--check_schema_path", type=str, help="Validation schema path.")
    parser.add_argument("--export_date", type=str, help="Export date in YYYY-MM format (for pipeline).")
    parser.add_argument("--field", type=str, help="Field name for add_canonical (document_type or issuing_party).")
    parser.add_argument("--canonical", type=str, help="Canonical value to add.")
    args = parser.parse_args()

    # Initialize logging early (reconfigures the module-level logger)
    setup_logging(verbose=args.verbose)

    # Initialize configuration (profile or legacy .env)
    try:
        initialize_config(args.profile)
    except ProfileNotFoundError as e:
        parser.error(str(e))
    except ProfileError as e:
        parser.error(f"Failed to load profile: {e}")

    if args.task == "pipeline":
        if args.export_date and not re.match(r"^\d{4}-\d{2}$", args.export_date):
            parser.error("The --export_date argument must be in YYYY-MM format.")
        pipeline(export_date_arg=args.export_date)
        return

    if args.task == "gmail_download":
        task_gmail_download()
        return

    if args.task == "review_mappings":
        task_review_mappings(get_ctx().mappings_manager)
        return

    if args.task == "add_canonical":
        if not args.field or not args.canonical:
            parser.error("add_canonical requires --field and --canonical arguments.")
        task_add_canonical(get_ctx().mappings_manager, args.field, args.canonical)
        return

    if args.task == "bootstrap_mappings":
        if not args.processed_path:
            parser.error("bootstrap_mappings requires the processed_path argument.")
        if not os.path.exists(args.processed_path):
            parser.error(f"The processed_path '{args.processed_path}' does not exist.")
        task_bootstrap_mappings(Path(args.processed_path), get_ctx().mappings_manager)
        return

    if not args.processed_path:
        parser.error("the processed_path argument is required.")
    if not os.path.exists(args.processed_path):
        parser.error(f"The processed_path '{args.processed_path}' does not exist.")
    if not os.path.isdir(args.processed_path):
        parser.error(f"The processed_path '{args.processed_path}' is not a directory.")

    raw_paths = None
    if args.task == "extract_new":
        if not args.raw_path:
            parser.error("the --raw_path argument is required when task is 'extract_new'.")
        raw_paths = [p for p in args.raw_path.split(';') if p]
        if not raw_paths:
            parser.error("the --raw_path argument must contain at least one path.")
        for rp in raw_paths:
            if not os.path.exists(rp):
                parser.error(f"The raw_path '{rp}' does not exist.")
            if not os.path.isdir(rp):
                parser.error(f"The raw_path '{rp}' is not a directory.")

    if args.task == "export_excel":
        if not args.excel_output_path:
            parser.error("the --excel_output_path argument is required when task is 'export_excel'.")
        if not args.excel_output_path.endswith(".xlsx"):
            parser.error("the --excel_output_path argument must end with '.xlsx'.")

    if args.task == "copy_matching":
        if not args.regex_pattern:
            parser.error("the --regex_pattern argument is required when task is 'copy_matching'.")
        if not args.copy_dest_folder:
            parser.error("the --copy_dest_folder argument is required when task is 'copy_matching'.")
        if not os.path.exists(args.copy_dest_folder):
            os.makedirs(args.copy_dest_folder, exist_ok=True)
        if not os.path.isdir(args.copy_dest_folder):
            parser.error(f"The copy_dest_folder '{args.copy_dest_folder}' is not a directory.")

    export_base_dir = args.export_base_dir
    if args.task == "export_all_dates":
        if not export_base_dir:
            export_base_dir = os.getenv("EXPORT_FILES_DIR")
            if not export_base_dir:
                parser.error("the --export_base_dir argument is required when task is 'export_all_dates'.")
        if not os.path.exists(export_base_dir):
            os.makedirs(export_base_dir, exist_ok=True)
        if not os.path.isdir(export_base_dir):
            parser.error(f"The export_base_dir '{export_base_dir}' is not a directory.")

    check_schema_path = args.check_schema_path
    temp_check_schema_file = None
    if args.task == "check_files_exist":
        if not check_schema_path:
            # Get validations from profile
            from papertrail.config import get_validations
            validations, validations_file = get_validations()
            if validations and validations.get('rules'):
                if validations_file:
                    check_schema_path = validations_file
                else:
                    # Create temp file for inline data
                    temp_check_schema = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json')
                    json.dump(validations['rules'], temp_check_schema, indent=2)
                    temp_check_schema.close()
                    check_schema_path = temp_check_schema.name
                    temp_check_schema_file = temp_check_schema.name
            else:
                parser.error("No validation rules found in profile. Use --check_schema_path or configure validations in profile.")
        if not os.path.exists(check_schema_path):
            parser.error(f"The check_schema_path '{check_schema_path}' does not exist.")

    process_folder(
        args.task,
        args.processed_path,
        raw_paths=raw_paths if args.task == "extract_new" else None,
        excel_output_path=args.excel_output_path,
        regex_pattern=args.regex_pattern,
        copy_dest_folder=args.copy_dest_folder,
        export_base_dir=export_base_dir,
        run_merge=args.run_merge if hasattr(args, 'run_merge') else False,
        check_schema_path=check_schema_path if args.task == "check_files_exist" else args.check_schema_path
    )

    # Cleanup temp schema file if created
    if temp_check_schema_file:
        try:
            os.unlink(temp_check_schema_file)
        except Exception:
            pass


if __name__ == "__main__":
    main()
